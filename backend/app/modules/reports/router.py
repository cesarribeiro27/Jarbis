"""
Endpoints de Relatórios — criação, consulta e compartilhamento de relatórios.

POST   /reports                         — cria relatório
GET    /reports                         — lista relatórios do tenant
GET    /reports/{id}                    — detalhe do relatório
PUT    /reports/{id}                    — atualiza relatório
DELETE /reports/{id}                    — remove relatório
POST   /reports/{id}/share              — gera link de compartilhamento
GET    /reports/public/{token}          — acesso público via token
GET    /reports/data/{source}           — dados de uma fonte pré-definida (legado)

GET    /reports/datasets                — lista datasets do tenant
POST   /reports/datasets/upload         — upload CSV/Excel
POST   /reports/datasets/api            — cria dataset via API
POST   /reports/datasets/{id}/sync      — re-sincroniza dataset de API
DELETE /reports/datasets/{id}           — remove dataset
GET    /reports/datasets/{id}/query     — consulta dataset com agregação
"""

import hashlib
import hmac
import json
import secrets
import uuid
from datetime import datetime, timezone
from typing import Annotated

import httpx
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.cache import cache_get, cache_set, get_redis, _sanitize_for_json
from app.database import get_db
from app.modules.auth.dependencies import get_current_active_user, get_effective_tenant_id
from app.modules.tenants.models import User

from app.modules.billing.guards import (
    check_ai_query_limit,
    check_alert_limit,
    check_dashboard_limit,
    check_dataset_limit,
    check_feature_allowed,
)
from app.modules.tenants.models import Tenant

from .dataset_models import ReportDataset
from .dataset_service import DatasetService
from .query_log_models import QueryLog
from .warp_cache import warp_get_rows, warp_set_rows, warp_invalidate, warp_status as _warp_status
from .schemas import (
    ReportCreate,
    ReportResponse,
    ReportSummary,
    ReportUpdate,
    ShareResponse,
)
from .service import ReportService


# ---------------------------------------------------------------------------
# Dataset schemas (inline para evitar arquivo extra)
# ---------------------------------------------------------------------------

class DatasetSummary(BaseModel):
    id: uuid.UUID
    name: str
    type: str
    columns: list[str]
    row_count: int
    column_types: dict | None = None
    is_demo: bool = False
    api_url: str | None = None
    last_synced_at: str | None = None
    refresh_interval_minutes: int | None = None
    next_refresh_at: str | None = None
    computed_columns: list[dict] = []

    model_config = {"from_attributes": True}


class ComputedColumnCreate(BaseModel):
    name: str
    expression: str
    refs: list[str] = []


class ComputedColumnResponse(BaseModel):
    name: str
    expression: str
    refs: list[str] = []


class ApiDatasetCreate(BaseModel):
    name: str
    api_url: str
    method: str = "GET"
    headers: dict | None = None          # frontend envia como "headers"
    api_headers: dict | None = None      # alias de compatibilidade
    body: str | None = None
    api_data_path: str | None = None
    refresh_interval_minutes: int | None = None
    sync_mode: str = "replace"           # replace | append

    @property
    def resolved_headers(self) -> dict | None:
        return self.headers or self.api_headers


class DbDatasetCreate(BaseModel):
    name: str
    db_type: str  # postgresql | mysql
    host: str
    port: int = 5432
    database: str
    username: str
    password: str
    query: str

router = APIRouter(prefix="/reports", tags=["Reports"])

_CACHE_TTL = 300  # 5 minutos


def _make_cache_key(dataset_id: str, query_params: dict, version: str = "") -> str:
    """Gera chave MD5 determinística para cache de query de dataset.

    O campo `version` permite invalidar todo o cache de um dataset de uma vez:
    basta mudar a versão armazenada em Redis (chave jarbis:ds_version:{dataset_id}).
    """
    params_str = json.dumps(query_params, sort_keys=True, default=str)
    h = hashlib.md5(f"query:{dataset_id}:{version}:{params_str}".encode()).hexdigest()
    return f"jarbis:query:{h}"


async def _get_dataset_version(redis, dataset_id: str) -> str:
    """Retorna a versão atual do dataset (usada para invalidação de cache)."""
    key = f"jarbis:ds_version:{dataset_id}"
    version = await redis.get(key)
    return version or "0"


async def _bump_dataset_version(redis, dataset_id: str) -> None:
    """Incrementa a versão do dataset, invalidando todo o cache de queries associado."""
    key = f"jarbis:ds_version:{dataset_id}"
    await redis.incr(key)
    # Mantém a chave de versão por 24h para não acumular entradas obsoletas
    await redis.expire(key, 86400)


# ---------------------------------------------------------------------------
# Preview anônimo (sem auth) — para mecânica de conversão da landing page
# ---------------------------------------------------------------------------

_PREVIEW_TTL = 86400  # 24 horas


@router.post(
    "/preview-upload",
    summary="Upload anônimo para preview — sem autenticação",
    include_in_schema=False,
)
async def preview_upload(
    file: Annotated[UploadFile, File()],
    request: Request,
):
    """Recebe arquivo CSV/Excel sem autenticação.
    Analisa, gera estatísticas e armazena em Redis com TTL 24h.
    Retorna temp_token para redirecionar para /preview/{token}.
    """
    from app.modules.reports.dataset_service import _parse_csv, _parse_excel, _detect_columns

    filename = file.filename or "planilha"
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande (máx 5 MB)")

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    try:
        if ext in ("xlsx", "xls"):
            rows = _parse_excel(content)
        else:
            rows = _parse_csv(content)
    except Exception:
        raise HTTPException(status_code=400, detail="Formato inválido. Use CSV ou Excel (.xlsx/.xls)")

    if not rows:
        raise HTTPException(status_code=400, detail="Arquivo vazio ou sem dados")

    columns = _detect_columns(rows)
    row_count = len(rows)

    # Gerar estatísticas simples por coluna
    stats = []
    for col in columns[:6]:  # máx 6 colunas para o preview
        vals = [r.get(col) for r in rows if r.get(col) is not None and str(r.get(col)).strip().lower() not in (
                "", "null", "none", "n/a", "na", "n.a.", "-", "--", "nan",
                "#n/a", "#na", "#div/0!", "#ref!", "#value!", "#num!", "#name?", "#null!",
            )]
        numeric = [v for v in vals if isinstance(v, (int, float))]
        if numeric:
            stats.append({
                "col": col,
                "type": "numeric",
                "sum": round(sum(numeric), 2),
                "avg": round(sum(numeric) / len(numeric), 2),
                "max": max(numeric),
                "min": min(numeric),
                "count": len(numeric),
            })
        else:
            unique = list(dict.fromkeys(str(v) for v in vals[:20]))
            stats.append({
                "col": col,
                "type": "text",
                "unique_count": len(set(str(v) for v in vals)),
                "top": unique[:5],
            })

    temp_token = secrets.token_urlsafe(24)
    preview_data = {
        "file_name": filename,
        "row_count": row_count,
        "columns": columns,
        "stats": stats,
        "sample": rows[:3],  # 3 linhas de amostra (não mostrar dados reais)
    }

    redis = await get_redis()
    if redis:
        import json as _json
        await redis.setex(
            f"jarbis:preview:{temp_token}",
            _PREVIEW_TTL,
            _json.dumps(preview_data, default=str),
        )

    return {"temp_token": temp_token, "row_count": row_count, "columns": columns}


@router.get(
    "/preview/{token}",
    summary="Retorna dados do preview anônimo",
    include_in_schema=False,
)
async def get_preview(token: str):
    """Retorna dados do preview sem autenticação."""
    import json as _json
    redis = await get_redis()
    if not redis:
        raise HTTPException(status_code=503, detail="Serviço indisponível")
    raw = await redis.get(f"jarbis:preview:{token}")
    if not raw:
        raise HTTPException(status_code=404, detail="Preview não encontrado ou expirado")
    return _json.loads(raw)


# ---------------------------------------------------------------------------
# Authenticated endpoints
# ---------------------------------------------------------------------------


@router.post(
    "",
    response_model=ReportResponse,
    status_code=201,
    summary="Cria um novo relatório",
)
async def create_report(
    data: ReportCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    await check_dashboard_limit(db, current_user.tenant_id, tenant.plan if tenant else "free", tenant.addon_packs if tenant else 0)
    service = ReportService(db)
    report = await service.create(current_user.tenant_id, data)
    return report


@router.get(
    "",
    response_model=list[ReportSummary],
    summary="Lista relatórios do tenant",
)
async def list_reports(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = ReportService(db)
    reports = await service.list(current_user.tenant_id)
    return [ReportSummary(**r) for r in reports]


@router.get(
    "/data/{source}",
    summary="Dados de uma fonte para widgets",
)
async def get_data_source(
    source: str,
    date_from: str | None = None,
    date_to: str | None = None,
    meio: str | None = None,
    anunciante: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Retorna dados no formato [{label, value}] prontos para uso em widgets de gráfico.

    Fontes disponíveis: items_by_meio, investment_by_meio, investment_by_client,
    investment_by_month, top_vehicles, material_by_status, pi_by_status,
    items_by_uf, vehicles_by_type, vehicles_by_uf, vehicles_by_status,
    municipios_by_regiao.
    """
    filters = {}
    if date_from:
        filters["date_from"] = date_from
    if date_to:
        filters["date_to"] = date_to
    if meio:
        filters["meio"] = meio
    if anunciante:
        filters["anunciante"] = anunciante

    service = ReportService(db)
    items = await service.get_data_source(source, current_user.tenant_id, filters=filters)
    return items


# ---------------------------------------------------------------------------
# Public endpoint (no auth) — MUST come before /{report_id} to avoid capture
# ---------------------------------------------------------------------------


@router.get(
    "/public/{token}",
    response_model=ReportResponse,
    summary="Acesso público ao relatório via token (sem autenticação)",
)
async def get_public_report(token: str, db: AsyncSession = Depends(get_db)):
    """
    Endpoint público acessado pelo link compartilhado.
    Incrementa o contador de visualizações a cada acesso.
    """
    service = ReportService(db)
    report = await service.get_public(token)
    if not report:
        raise HTTPException(status_code=404, detail="Link inválido ou expirado")
    tenant = await db.scalar(select(Tenant).where(Tenant.id == report.tenant_id))
    check_feature_allowed("embed", tenant.plan if tenant else "free")
    return report


@router.get(
    "/public/{token}/datasets/{dataset_id}/query",
    summary="Query pública de dataset via share token (sem autenticação)",
)
async def public_query_dataset(
    token: str,
    dataset_id: uuid.UUID,
    label_col: str,
    value_col: str,
    agg: str = "sum",
    filter_col: str | None = None,
    filter_val: str | None = None,
    date_col: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    """Permite que páginas públicas (embed/share) consultem datasets sem JWT."""
    report_service = ReportService(db)
    report = await report_service.get_public_no_increment(token)
    if not report:
        raise HTTPException(status_code=404, detail="Link inválido ou expirado")

    query_params = {
        "label_col": label_col, "value_col": value_col, "agg": agg,
        "filter_col": filter_col, "filter_val": filter_val,
        "date_col": date_col, "date_from": date_from, "date_to": date_to,
        "tenant_id": str(report.tenant_id),
    }

    redis = get_redis()
    try:
        version = await _get_dataset_version(redis, str(dataset_id))
        cache_key = _make_cache_key(str(dataset_id), query_params, version)

        # 1. Check query cache (sem DB)
        cached = await cache_get(redis, cache_key)
        if cached is not None:
            return JSONResponse(content=cached, headers={"Cache-Control": f"public, max-age={_CACHE_TTL}"})

        # 2. Full DB load (popula Warp para próximas queries)
        dataset_service = DatasetService(db)
        ds = await dataset_service.get(dataset_id, report.tenant_id)
        if not ds:
            raise HTTPException(status_code=404, detail="Dataset não encontrado")

        from .dataset_service import _apply_computed_columns
        computed = getattr(ds, 'computed_columns', None) or []
        warp_rows = _apply_computed_columns(ds.rows or [], computed) if computed else (ds.rows or [])
        await warp_set_rows(redis, str(dataset_id), warp_rows)

        result = dataset_service.query(ds, label_col, value_col, agg, filter_col, filter_val, date_col, date_from, date_to)
        await cache_set(redis, cache_key, result, ttl=_CACHE_TTL)
        return JSONResponse(content=result, headers={"Cache-Control": f"public, max-age={_CACHE_TTL}"})
    finally:
        await redis.aclose()


# ---------------------------------------------------------------------------
# Query Logs
# ---------------------------------------------------------------------------


@router.get("/query-logs")
async def list_query_logs(
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_active_user),
):
    result = await db.execute(
        select(QueryLog)
        .where(QueryLog.tenant_id == user.tenant_id)
        .order_by(QueryLog.created_at.desc())
        .limit(limit)
    )
    logs = result.scalars().all()
    return [
        {
            "id": l.id,
            "dataset_id": str(l.dataset_id) if l.dataset_id else None,
            "dataset_name": l.dataset_name,
            "query_type": l.query_type,
            "duration_ms": l.duration_ms,
            "rows_returned": l.rows_returned,
            "status": l.status,
            "error_msg": l.error_msg,
            "cached": l.cached,
            "created_at": l.created_at.isoformat() if l.created_at else None,
        }
        for l in logs
    ]


# ---------------------------------------------------------------------------
# Datasets — fontes de dados gerenciadas pelo usuário
# ---------------------------------------------------------------------------


@router.get(
    "/datasets",
    response_model=list[DatasetSummary],
    summary="Lista datasets do tenant",
)
async def list_datasets(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    effective_tenant_id: uuid.UUID = Depends(get_effective_tenant_id),
):
    service = DatasetService(db)
    # Se em contexto de sub-org, une datasets da sub-org + datasets do tenant raiz (compartilhados)
    if effective_tenant_id != current_user.tenant_id:
        suborg_datasets = await service.list(effective_tenant_id)
        root_datasets = await service.list(current_user.tenant_id)
        seen = set()
        datasets = []
        for ds in suborg_datasets + root_datasets:
            if ds.id not in seen:
                seen.add(ds.id)
                datasets.append(ds)
    else:
        datasets = await service.list(current_user.tenant_id)
    from .query_engine import detect_column_types
    return [
        DatasetSummary(
            id=ds.id,
            name=ds.name,
            type=ds.type,
            columns=ds.columns or [],
            row_count=ds.row_count,
            column_types=detect_column_types(ds.rows or [], sample_size=30) if ds.rows else {},
            is_demo=bool(ds.is_demo),
            api_url=ds.api_url,
            last_synced_at=ds.last_synced_at.isoformat() if ds.last_synced_at else None,
            refresh_interval_minutes=ds.refresh_interval_minutes,
            next_refresh_at=ds.next_refresh_at.isoformat() if ds.next_refresh_at else None,
            computed_columns=ds.computed_columns or [],
        )
        for ds in datasets
    ]


@router.get(
    "/onboarding-dataset",
    response_model=DatasetSummary,
    summary="Retorna (ou cria) o dataset de demonstração do tenant",
)
async def get_onboarding_dataset(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    from .onboarding import ensure_onboarding_dataset
    from .query_engine import detect_column_types
    ds = await ensure_onboarding_dataset(current_user.tenant_id, db)
    return DatasetSummary(
        id=ds.id,
        name=ds.name,
        type=ds.type,
        columns=ds.columns or [],
        row_count=ds.row_count,
        column_types=detect_column_types(ds.rows or [], sample_size=30) if ds.rows else {},
        is_demo=True,
    )


def _analyze_sheet(ws) -> dict:
    """Analisa uma aba e retorna metadados para ajudar o usuário a identificar bancos de dados."""
    import re as _re

    # Nomes suspeitos de não serem banco de dados
    SUMMARY_KEYWORDS = {
        'dashboard', 'analise', 'análise', 'visao', 'visão', 'resumo',
        'pivot', 'grafico', 'gráfico', 'relatorio', 'relatório', 'chart',
        'overview', 'sumario', 'sumário', 'summary', 'report',
    }

    name_lower = ws.title.lower()
    name_clean = _re.sub(r'[^a-z]', '', name_lower)

    # Lê amostra de linhas (máx 100)
    sample_rows = []
    for row in ws.iter_rows(values_only=True, max_row=101):
        sample_rows.append(row)
        if len(sample_rows) >= 101:
            break

    if not sample_rows:
        return {"name": ws.title, "type": "empty", "row_count": 0, "col_count": 0, "suggested": False, "reason": "Aba vazia"}

    from app.modules.reports.dataset_service import _find_header_row
    header_idx = _find_header_row(sample_rows)
    headers = sample_rows[header_idx]
    data_rows = [r for r in sample_rows[header_idx + 1:] if any(v is not None for v in r)]
    col_count = sum(1 for h in headers if h is not None)
    row_count = len(data_rows)  # approximate (sample only)

    # Heurísticas
    name_is_summary = any(kw in name_clean for kw in SUMMARY_KEYWORDS)
    has_enough_rows = row_count >= 5
    has_enough_cols = col_count >= 3
    headers_look_like_text = sum(1 for h in headers if h and isinstance(h, str) and not str(h).replace('.','').replace(',','').replace(' ','').replace('R$','').replace('%','').lstrip('-').isnumeric()) >= col_count * 0.6 if col_count else False

    # Fill density: percentual de células não-nulas nas linhas de dados
    if data_rows and col_count:
        filled = sum(1 for r in data_rows[:20] for v in r[:col_count] if v is not None)
        total = len(data_rows[:20]) * col_count
        fill_density = filled / total if total else 0
    else:
        fill_density = 0

    # Score: 0-100
    score = 0
    if has_enough_rows: score += 30
    if has_enough_cols: score += 20
    if headers_look_like_text: score += 20
    if fill_density >= 0.5: score += 20
    if not name_is_summary: score += 10

    if score >= 60 and has_enough_rows and has_enough_cols:
        sheet_type = "data"
        suggested = True
        reason = f"{row_count}+ linhas · {col_count} colunas · parece banco de dados"
    elif row_count == 0:
        sheet_type = "empty"
        suggested = False
        reason = "Aba vazia ou sem dados"
    elif name_is_summary or not has_enough_rows:
        sheet_type = "summary"
        suggested = False
        reason = "Parece resumo ou tabela analítica"
    else:
        sheet_type = "unknown"
        suggested = False
        reason = f"{row_count} linhas · estrutura indefinida"

    return {
        "name": ws.title,
        "type": sheet_type,       # "data" | "summary" | "empty" | "unknown"
        "row_count": row_count,
        "col_count": col_count,
        "suggested": suggested,
        "reason": reason,
    }


def _enrich_formula_mirrors(content: bytes, sheets_meta: list[dict]) -> None:
    """Detecta abas que são espelhos por fórmula (sem dados cached) e enriquece os metadados.

    Modifica sheets_meta in-place: abas-espelho recebem type='formula',
    formula_source com o nome da aba real, e suggested=False.
    A aba fonte tem seu suggested=True forçado se ainda não estava.
    """
    from app.modules.reports.dataset_service import _resolve_formula_sheet
    for meta in sheets_meta:
        if meta.get("row_count", 0) < 3 and meta.get("type") in ("empty", "unknown", "summary"):
            source = _resolve_formula_sheet(content, meta["name"])
            if source:
                meta["type"] = "formula"
                meta["formula_source"] = source
                meta["suggested"] = False
                meta["reason"] = f"Espelho de '{source}' — importe a aba '{source}'"
                # Garante que a aba fonte seja marcada como sugerida
                for m in sheets_meta:
                    if m["name"] == source and not m.get("suggested"):
                        m["suggested"] = True


@router.get(
    "/datasets/google-sheets-sheets",
    summary="Lista abas de uma planilha Google Sheets pública",
)
async def get_google_sheets_sheets(
    url: str = Query(..., description="URL pública do Google Sheets"),
    current_user: User = Depends(get_current_active_user),
):
    import re
    m = re.search(r"spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not m:
        raise HTTPException(status_code=400, detail="URL do Google Sheets inválida")
    spreadsheet_id = m.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    try:
        async with httpx.AsyncClient(follow_redirects=True, timeout=15) as client:
            resp = await client.get(export_url)
        if resp.status_code != 200:
            raise HTTPException(status_code=400, detail="Não foi possível baixar a planilha. Verifique se está pública.")
        import io
        import openpyxl
        content = resp.content
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets_meta = [_analyze_sheet(wb[name]) for name in wb.sheetnames]
        wb.close()
        _enrich_formula_mirrors(content, sheets_meta)
        return {"sheets": [s["name"] for s in sheets_meta], "sheets_meta": sheets_meta}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler planilha: {e}")


@router.post(
    "/datasets/excel-sheets",
    summary="Lista abas de um arquivo Excel sem criar dataset",
)
async def get_excel_sheets(
    file: Annotated[UploadFile, File()],
    current_user: User = Depends(get_current_active_user),
):
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande (máx 20 MB)")
    try:
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets_meta = [_analyze_sheet(wb[name]) for name in wb.sheetnames]
        wb.close()
        _enrich_formula_mirrors(content, sheets_meta)
        return {"sheets": [s["name"] for s in sheets_meta], "sheets_meta": sheets_meta}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler arquivo: {e}")


@router.post(
    "/datasets/upload",
    response_model=DatasetSummary,
    status_code=201,
    summary="Upload CSV ou Excel para criar dataset",
)
async def upload_dataset(
    file: Annotated[UploadFile, File()],
    sheet_name: str | None = Form(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    effective_tenant_id: uuid.UUID = Depends(get_effective_tenant_id),
):
    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    await check_dataset_limit(db, current_user.tenant_id, tenant.plan if tenant else "free", tenant.addon_packs if tenant else 0)
    filename = file.filename or "dataset"
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande (máx 20 MB)")
    name = filename.rsplit(".", 1)[0] if "." in filename else filename
    service = DatasetService(db)
    ds = await service.create_from_file(effective_tenant_id, name, filename, content, sheet_name=sheet_name)
    from .query_engine import detect_column_types
    return DatasetSummary(
        id=ds.id, name=ds.name, type=ds.type,
        columns=ds.columns or [], row_count=ds.row_count,
        column_types=detect_column_types(ds.rows or [], sample_size=30) if ds.rows else {},
    )


@router.post(
    "/datasets/api",
    response_model=DatasetSummary,
    status_code=201,
    summary="Cria dataset a partir de uma API externa",
)
async def create_api_dataset(
    body: ApiDatasetCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    effective_tenant_id: uuid.UUID = Depends(get_effective_tenant_id),
):
    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    await check_dataset_limit(db, current_user.tenant_id, tenant.plan if tenant else "free", tenant.addon_packs if tenant else 0)
    service = DatasetService(db)
    try:
        ds = await service.create_from_api(
            effective_tenant_id,
            body.name,
            body.api_url,
            body.resolved_headers,
            body.api_data_path,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao buscar API: {e}")

    # Salva sync_mode e refresh schedule se informados
    ds.sync_mode = body.sync_mode or "replace"
    if body.refresh_interval_minutes:
        from datetime import timedelta
        ds.refresh_interval_minutes = body.refresh_interval_minutes
        ds.next_refresh_at = datetime.now(timezone.utc) + timedelta(minutes=body.refresh_interval_minutes)
    await db.commit()
    await db.refresh(ds)

    return DatasetSummary(
        id=ds.id, name=ds.name, type=ds.type,
        columns=ds.columns or [], row_count=ds.row_count,
        api_url=ds.api_url,
        last_synced_at=ds.last_synced_at.isoformat() if ds.last_synced_at else None,
        refresh_interval_minutes=ds.refresh_interval_minutes,
    )


@router.post(
    "/datasets/database/test",
    summary="Testa conexão com banco de dados externo",
)
async def test_db_connection(
    body: DbDatasetCreate,
    current_user=Depends(get_current_active_user),
):
    from .connectors.db_connector import test_connection
    result = await test_connection(body.db_type, body.host, body.port, body.database, body.username, body.password)
    return result


@router.post(
    "/datasets/database",
    status_code=201,
    summary="Cria dataset a partir de banco de dados externo (PostgreSQL/MySQL)",
)
async def create_db_dataset(
    body: DbDatasetCreate,
    current_user=Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    await check_dataset_limit(db, current_user.tenant_id, tenant.plan if tenant else "free", tenant.addon_packs if tenant else 0)
    from .connectors.db_connector import encrypt_password, run_query
    rows = await run_query(body.db_type, body.host, body.port, body.database, body.username, body.password, body.query)

    from .dataset_service import _detect_columns, _coerce_row
    coerced = [_coerce_row(r) for r in rows]
    columns = _detect_columns(coerced)

    ds = ReportDataset(
        tenant_id=current_user.tenant_id,
        name=body.name,
        type="database",
        rows=coerced,
        columns=columns,
        row_count=len(coerced),
        db_type=body.db_type,
        db_host=body.host,
        db_port=body.port,
        db_name=body.database,
        db_username=body.username,
        db_password_enc=encrypt_password(body.password),
        db_query=body.query,
    )
    db.add(ds)
    await db.commit()
    await db.refresh(ds)
    return {"id": str(ds.id), "name": ds.name, "row_count": ds.row_count, "columns": ds.columns}


@router.post(
    "/datasets/{dataset_id}/database/sync",
    summary="Re-sincroniza dataset de banco de dados externo",
)
async def sync_db_dataset(
    dataset_id: uuid.UUID,
    current_user=Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy.orm.attributes import flag_modified
    from .connectors.db_connector import decrypt_password, run_query
    from .dataset_service import _detect_columns, _coerce_row

    ds = await db.scalar(select(ReportDataset).where(
        ReportDataset.id == dataset_id, ReportDataset.tenant_id == current_user.tenant_id
    ))
    if not ds or ds.type != "database":
        raise HTTPException(status_code=404, detail="Dataset de banco não encontrado")

    password = decrypt_password(ds.db_password_enc)
    rows = await run_query(ds.db_type, ds.db_host, ds.db_port, ds.db_name, ds.db_username, password, ds.db_query)
    coerced = [_coerce_row(r) for r in rows]

    ds.rows = coerced
    ds.columns = _detect_columns(coerced)
    ds.row_count = len(coerced)
    flag_modified(ds, 'rows')
    flag_modified(ds, 'columns')
    await db.commit()

    # Invalida Warp + versão de cache após sincronização
    try:
        redis = get_redis()
        try:
            await warp_invalidate(redis, str(dataset_id))
            await _bump_dataset_version(redis, str(dataset_id))
        finally:
            await redis.aclose()
    except Exception:
        pass

    return {"ok": True, "row_count": ds.row_count}


@router.post(
    "/datasets/{dataset_id}/sync",
    response_model=DatasetSummary,
    summary="Re-sincroniza dataset de API",
)
async def sync_dataset(
    dataset_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = DatasetService(db)
    try:
        ds = await service.sync_api(dataset_id, current_user.tenant_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao sincronizar: {e}")
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado ou não é do tipo API")

    # Invalida Warp + cache de queries deste dataset incrementando sua versão
    try:
        redis = get_redis()
        try:
            await warp_invalidate(redis, str(dataset_id))
            await _bump_dataset_version(redis, str(dataset_id))
        finally:
            await redis.aclose()
    except Exception:
        pass  # Falha no cache não deve bloquear a resposta

    return DatasetSummary(
        id=ds.id, name=ds.name, type=ds.type,
        columns=ds.columns or [], row_count=ds.row_count,
        api_url=ds.api_url,
        last_synced_at=ds.last_synced_at.isoformat() if ds.last_synced_at else None,
    )


class DatasetScheduleUpdate(BaseModel):
    refresh_interval_minutes: int | None  # None = disable


@router.patch("/datasets/{dataset_id}/schedule", summary="Configura refresh automático do dataset")
async def set_dataset_schedule(
    dataset_id: uuid.UUID,
    data: DatasetScheduleUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    from datetime import datetime as _dt, timedelta as _td, timezone as _tz2
    svc = DatasetService(db)
    ds = await svc.get(dataset_id, current_user.tenant_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")
    ds.refresh_interval_minutes = data.refresh_interval_minutes
    if data.refresh_interval_minutes:
        ds.next_refresh_at = _dt.now(_tz2.utc) + _td(minutes=data.refresh_interval_minutes)
    else:
        ds.next_refresh_at = None
    await db.commit()
    await db.refresh(ds)
    return {
        "id": str(ds.id),
        "refresh_interval_minutes": ds.refresh_interval_minutes,
        "next_refresh_at": ds.next_refresh_at.isoformat() if ds.next_refresh_at else None,
    }


@router.delete("/datasets/{dataset_id}", status_code=204, summary="Remove dataset")
async def delete_dataset(
    dataset_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = DatasetService(db)
    deleted = await service.delete(dataset_id, current_user.tenant_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")


@router.get(
    "/datasets/{dataset_id}/query",
    summary="Consulta dataset com agregação — retorna [{label, value}]",
)
async def query_dataset(
    dataset_id: uuid.UUID,
    label_col: str,
    value_col: str,
    agg: str = "sum",
    filter_col: str | None = None,
    filter_val: str | None = None,
    date_col: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    effective_tenant_id: uuid.UUID = Depends(get_effective_tenant_id),
):
    query_params = {
        "label_col": label_col, "value_col": value_col, "agg": agg,
        "filter_col": filter_col, "filter_val": filter_val,
        "date_col": date_col, "date_from": date_from, "date_to": date_to,
        "tenant_id": str(effective_tenant_id),
    }

    redis = get_redis()
    try:
        version = await _get_dataset_version(redis, str(dataset_id))
        cache_key = _make_cache_key(str(dataset_id), query_params, version)

        # 1. Check query cache (sem DB)
        cached = await cache_get(redis, cache_key)
        if cached is not None:
            return cached

        # 2. Carrega dataset do DB (fallback para tenant raiz se suborg não tiver o dataset)
        service = DatasetService(db)
        ds = await service.get(dataset_id, effective_tenant_id)
        if not ds and effective_tenant_id != current_user.tenant_id:
            ds = await service.get(dataset_id, current_user.tenant_id)
        if not ds:
            raise HTTPException(status_code=404, detail="Dataset não encontrado")

        # 3. Popula Warp com rows pré-computadas (inclui computed_columns)
        from .dataset_service import _apply_computed_columns
        computed = getattr(ds, 'computed_columns', None) or []
        warp_rows = _apply_computed_columns(ds.rows or [], computed) if computed else (ds.rows or [])
        await warp_set_rows(redis, str(dataset_id), warp_rows)

        result = service.query(ds, label_col, value_col, agg, filter_col, filter_val, date_col, date_from, date_to)
        await cache_set(redis, cache_key, result, ttl=_CACHE_TTL)
        return result
    finally:
        await redis.aclose()


# ---------------------------------------------------------------------------
# Query Engine v2 — motor de query estruturado
# ---------------------------------------------------------------------------

from .query_engine import QueryRequest, QueryResult, detect_column_types


@router.post(
    "/datasets/{dataset_id}/query",
    response_model=QueryResult,
    summary="Consulta dataset com motor de query estruturado (v2)",
)
async def query_dataset_v2(
    dataset_id: uuid.UUID,
    req: QueryRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    effective_tenant_id: uuid.UUID = Depends(get_effective_tenant_id),
):
    """
    Motor de query v2. Aceita dimensões, métricas, filtros, date_range,
    ordenação e limite. Retorna dados normalizados para qualquer visualização.
    """
    import logging as _logging
    from .query_engine import execute_query
    query_params = {**req.model_dump(), "tenant_id": str(effective_tenant_id)}

    redis = get_redis()
    cache_key = None
    try:
        # Redis opcional — se indisponível, segue direto para DB
        try:
            version = await _get_dataset_version(redis, str(dataset_id))
            cache_key = _make_cache_key(str(dataset_id), query_params, version)

            # 1. Check query cache (sem DB)
            cached = await cache_get(redis, cache_key)
            if cached is not None:
                return cached

            # 2. Check Warp rows (evita ler JSONB pesado do DB)
            warp_rows = await warp_get_rows(redis, str(dataset_id))
            if warp_rows is not None:
                ds_exists = await db.scalar(
                    select(ReportDataset.id).where(
                        ReportDataset.id == dataset_id,
                        ReportDataset.tenant_id.in_(
                            [effective_tenant_id, current_user.tenant_id]
                        ),
                    )
                )
                if ds_exists:
                    try:
                        result = execute_query(warp_rows, req)
                        result_dict = _sanitize_for_json(result.model_dump() if hasattr(result, "model_dump") else result)
                        await cache_set(redis, cache_key, result_dict, ttl=_CACHE_TTL)
                        return result_dict
                    except Exception as exc:
                        _logging.getLogger(__name__).error("execute_query (warp) failed: %s", exc, exc_info=True)
                        return {"data": [], "total_rows": 0, "columns": [], "compare_data": None}
        except Exception as redis_exc:
            _logging.getLogger(__name__).warning("Redis unavailable, falling back to DB: %s", redis_exc)

        # 3. Full DB load (popula Warp para próximas queries); fallback para tenant raiz
        service = DatasetService(db)
        ds = await service.get(dataset_id, effective_tenant_id)
        if not ds and effective_tenant_id != current_user.tenant_id:
            ds = await service.get(dataset_id, current_user.tenant_id)
        if not ds:
            raise HTTPException(status_code=404, detail="Dataset não encontrado")

        from .dataset_service import _apply_computed_columns
        computed = getattr(ds, 'computed_columns', None) or []
        rows = _apply_computed_columns(ds.rows or [], computed) if computed else (ds.rows or [])
        try:
            await warp_set_rows(redis, str(dataset_id), rows)
        except Exception:
            pass  # Cache failure não deve derrubar a query

        try:
            result = execute_query(rows, req)
            result_dict = _sanitize_for_json(result.model_dump() if hasattr(result, "model_dump") else result)
            if cache_key:
                await cache_set(redis, cache_key, result_dict, ttl=_CACHE_TTL)
            return result_dict
        except Exception as exc:
            _logging.getLogger(__name__).error("execute_query failed: %s", exc, exc_info=True)
            return {"data": [], "total_rows": 0, "columns": [], "compare_data": None}
    finally:
        await redis.aclose()


@router.get(
    "/datasets/{dataset_id}/warp-status",
    summary="Status do cache Warp para o dataset",
)
async def get_warp_status(
    dataset_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = DatasetService(db)
    ds = await service.get(dataset_id, current_user.tenant_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")
    redis = get_redis()
    try:
        status = await _warp_status(redis, str(dataset_id))
        return status
    finally:
        await redis.aclose()


@router.post(
    "/public/{token}/datasets/{dataset_id}/query",
    response_model=QueryResult,
    summary="Query pública v2 via share token (sem autenticação)",
)
async def public_query_dataset_v2(
    token: str,
    dataset_id: uuid.UUID,
    req: QueryRequest,
    db: AsyncSession = Depends(get_db),
):
    import logging as _logging
    from .query_engine import execute_query
    report_service = ReportService(db)
    report = await report_service.get_public_no_increment(token)
    if not report:
        raise HTTPException(status_code=404, detail="Link inválido ou expirado")

    query_params = {**req.model_dump(), "tenant_id": str(report.tenant_id)}

    redis = get_redis()
    cache_key = None
    try:
        # Redis opcional — se indisponível, segue direto para DB
        try:
            version = await _get_dataset_version(redis, str(dataset_id))
            cache_key = _make_cache_key(str(dataset_id), query_params, version)

            # 1. Check query cache (sem DB)
            cached = await cache_get(redis, cache_key)
            if cached is not None:
                return JSONResponse(content=cached, headers={"Cache-Control": f"public, max-age={_CACHE_TTL}"})

            # 2. Check Warp rows (evita ler JSONB pesado do DB)
            warp_rows = await warp_get_rows(redis, str(dataset_id))
            if warp_rows is not None:
                ds_exists = await db.scalar(
                    select(ReportDataset.id).where(
                        ReportDataset.id == dataset_id,
                        ReportDataset.tenant_id == report.tenant_id,
                    )
                )
                if ds_exists:
                    result = execute_query(warp_rows, req)
                    result_dict = _sanitize_for_json(result.model_dump() if hasattr(result, "model_dump") else result)
                    await cache_set(redis, cache_key, result_dict, ttl=_CACHE_TTL)
                    return JSONResponse(
                        content=result_dict,
                        headers={"Cache-Control": f"public, max-age={_CACHE_TTL}"},
                    )
        except Exception as redis_exc:
            _logging.getLogger(__name__).warning("Redis unavailable (public query), falling back to DB: %s", redis_exc)

        # 3. Full DB load (popula Warp para próximas queries)
        dataset_service = DatasetService(db)
        ds = await dataset_service.get(dataset_id, report.tenant_id)
        if not ds:
            raise HTTPException(status_code=404, detail="Dataset não encontrado")

        from .dataset_service import _apply_computed_columns
        computed = getattr(ds, 'computed_columns', None) or []
        rows = _apply_computed_columns(ds.rows or [], computed) if computed else (ds.rows or [])
        try:
            await warp_set_rows(redis, str(dataset_id), rows)
        except Exception:
            pass

        result = execute_query(rows, req)
        result_dict = _sanitize_for_json(result.model_dump() if hasattr(result, "model_dump") else result)
        if cache_key:
            await cache_set(redis, cache_key, result_dict, ttl=_CACHE_TTL)
        return JSONResponse(
            content=result_dict,
            headers={"Cache-Control": f"public, max-age={_CACHE_TTL}"},
        )
    finally:
        await redis.aclose()


@router.get(
    "/datasets/{dataset_id}/columns",
    summary="Retorna colunas do dataset com tipos detectados automaticamente",
)
async def get_dataset_columns(
    dataset_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Retorna lista de colunas com tipo detectado: text | number | date.
    Usado pelo frontend para montar o painel de configuração de blocos.
    """
    service = DatasetService(db)
    ds = await service.get(dataset_id, current_user.tenant_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")
    types = detect_column_types(ds.rows or [])
    return {
        "dataset_id": str(dataset_id),
        "name": ds.name,
        "row_count": ds.row_count,
        "columns": [
            {"name": col, "type": types.get(col, "text")}
            for col in (ds.columns or [])
        ],
    }


# ---------------------------------------------------------------------------
# Dataset CRUD — rows, rename, delete column
# ---------------------------------------------------------------------------


class DatasetUpdateRequest(BaseModel):
    name: str


@router.get(
    "/datasets/{dataset_id}/rows",
    summary="Retorna linhas brutas do dataset com paginação",
)
async def get_dataset_rows(
    dataset_id: uuid.UUID,
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = DatasetService(db)
    ds = await service.get(dataset_id, current_user.tenant_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")
    rows = ds.rows or []
    return {"rows": rows[offset:offset + limit], "total": len(rows)}


@router.patch(
    "/datasets/{dataset_id}",
    response_model=DatasetSummary,
    summary="Renomeia um dataset",
)
async def update_dataset(
    dataset_id: uuid.UUID,
    body: DatasetUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = DatasetService(db)
    ds = await service.get(dataset_id, current_user.tenant_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")
    ds.name = body.name.strip()
    await db.commit()
    await db.refresh(ds)
    return DatasetSummary(
        id=ds.id, name=ds.name, type=ds.type,
        columns=ds.columns or [], row_count=ds.row_count or 0,
        api_url=ds.api_url,
        last_synced_at=ds.last_synced_at.isoformat() if ds.last_synced_at else None,
    )


@router.delete(
    "/datasets/{dataset_id}/columns/{column_name}",
    summary="Remove uma coluna do dataset",
)
async def delete_dataset_column(
    dataset_id: uuid.UUID,
    column_name: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    from sqlalchemy.orm.attributes import flag_modified
    service = DatasetService(db)
    ds = await service.get(dataset_id, current_user.tenant_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")
    if not ds.columns or column_name not in ds.columns:
        raise HTTPException(status_code=404, detail="Coluna não encontrada")
    ds.rows = [{k: v for k, v in row.items() if k != column_name} for row in (ds.rows or [])]
    ds.columns = [c for c in (ds.columns or []) if c != column_name]
    ds.row_count = len(ds.rows)
    flag_modified(ds, "rows")
    flag_modified(ds, "columns")
    await db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Computed Columns — colunas calculadas por expressão
# ---------------------------------------------------------------------------


@router.post(
    "/datasets/{dataset_id}/computed-columns",
    response_model=list[ComputedColumnResponse],
    summary="Adiciona ou atualiza uma coluna calculada no dataset",
)
async def add_computed_column(
    dataset_id: uuid.UUID,
    body: ComputedColumnCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = DatasetService(db)
    ds = await service.add_computed_column(
        dataset_id,
        current_user.tenant_id,
        body.name.strip(),
        body.expression.strip(),
        body.refs,
    )
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")
    return [
        ComputedColumnResponse(**c)
        for c in (ds.computed_columns or [])
    ]


@router.delete(
    "/datasets/{dataset_id}/computed-columns/{col_name}",
    response_model=list[ComputedColumnResponse],
    summary="Remove uma coluna calculada do dataset",
)
async def remove_computed_column(
    dataset_id: uuid.UUID,
    col_name: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = DatasetService(db)
    ds = await service.remove_computed_column(
        dataset_id,
        current_user.tenant_id,
        col_name,
    )
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")
    return [
        ComputedColumnResponse(**c)
        for c in (ds.computed_columns or [])
    ]


# ---------------------------------------------------------------------------
# AI — query em linguagem natural
# ---------------------------------------------------------------------------


def _suggest_chart(data: list[dict], question: str) -> dict:
    """Sugere tipo de gráfico baseado nos dados."""
    q_lower = question.lower()
    num_rows = len(data)

    # Detect column types
    cols = list(data[0].keys()) if data else []
    label_col = cols[0] if cols else None
    value_col = cols[1] if len(cols) > 1 else None

    chart_type = "bar"
    if any(w in q_lower for w in ["pizza", "pie", "percentual", "distribui"]):
        chart_type = "pie"
    elif any(w in q_lower for w in ["linha", "line", "tendência", "evolução", "histórico", "tempo"]):
        chart_type = "line"
    elif any(w in q_lower for w in ["funil", "funnel", "etapa", "conversão"]):
        chart_type = "funnel"
    elif any(w in q_lower for w in ["mapa", "estado", "região", "uf"]):
        chart_type = "map"
    elif num_rows <= 6:
        chart_type = "pie"

    return {
        "suggested_chart_type": chart_type,
        "suggested_title": question[:80],
        "suggested_config": {
            "label_col": label_col,
            "value_col": value_col,
            "chart_type": chart_type,
        },
    }


class AiQueryRequest(BaseModel):
    dataset_id: uuid.UUID
    question: str


_AI_MODEL = "claude-haiku-4-5-20251001"
_AI_MODEL_GENERATE = "claude-sonnet-4-6"


@router.post("/ai-query", summary="Consulta dataset com linguagem natural via IA (Claude)")
async def ai_query_endpoint(
    data: AiQueryRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    effective_tenant_id: uuid.UUID = Depends(get_effective_tenant_id),
):
    import json
    import os
    import re

    import anthropic as ant

    from .ai_usage_models import AIUsageLog

    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    plan = tenant.plan if tenant else "free"
    check_feature_allowed("ai", plan)
    await check_ai_query_limit(db, current_user.tenant_id, plan)

    # ── Verificação de segurança — bloqueia prompt injection e engenharia social ──
    from .security import check_question_safety
    safety = await check_question_safety(
        question=data.question,
        db=db,
        tenant_id=current_user.tenant_id,
        user_id=current_user.id,
        endpoint="ai_query",
    )
    if not safety["safe"]:
        raise HTTPException(status_code=403, detail={
            "code": "safety_violation",
            "message": safety["message"],
            "incident_count": safety.get("incident_count", 1),
        })

    svc = DatasetService(db)
    ds = await svc.get(data.dataset_id, effective_tenant_id)
    if not ds and effective_tenant_id != current_user.tenant_id:
        ds = await svc.get(data.dataset_id, current_user.tenant_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key or api_key == "your_key_here":
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY não configurada. Adicione a chave no arquivo .env e reinicie o backend.")

    columns = ds.columns or []
    rows = ds.rows or []
    total_rows = len(rows)
    sample = rows[:5]

    # ── Análise estatística completa de TODAS as linhas (igual ao generate_dashboard) ──
    col_stats: dict[str, dict] = {}
    for col in columns:
        vals = [r.get(col) for r in rows if r.get(col) is not None and str(r.get(col)).strip().lower() not in (
                "", "null", "none", "n/a", "na", "n.a.", "-", "--", "nan",
                "#n/a", "#na", "#div/0!", "#ref!", "#value!", "#num!", "#name?", "#null!",
            )]
        null_count = total_rows - len(vals)
        if not vals:
            col_stats[col] = {"type": "texto", "note": "sem dados"}
            continue
        nums: list[float] = []
        all_numeric = True
        for v in vals:
            try:
                nums.append(float(str(v).replace(",", ".")))
            except (ValueError, TypeError):
                all_numeric = False
                break
        if all_numeric and nums:
            nonzero = sum(1 for n in nums if n != 0)
            col_stats[col] = {
                "type": "número",
                "sum": round(sum(nums), 2),
                "avg": round(sum(nums) / len(nums), 2),
                "nonzero_pct": round(nonzero / len(nums) * 100),
                "null_pct": round(null_count / total_rows * 100) if total_rows else 0,
            }
        else:
            sample_val = str(vals[0])
            if re.match(r"\d{4}-\d{2}|\d{2}/\d{4}|\d{2}/\d{2}/\d{4}|\d{4}/\d{2}/\d{2}", sample_val):
                col_stats[col] = {"type": "data", "distinct": len(set(str(v) for v in vals))}
            else:
                uniq = list(dict.fromkeys(str(v) for v in vals))
                col_stats[col] = {"type": "texto", "distinct": len(uniq), "top3": uniq[:3]}

    # Montar descrição de colunas com estatísticas
    col_lines = []
    for col in columns:
        st = col_stats.get(col, {})
        t = st.get("type", "texto")
        if t == "número":
            null_pct = st.get("null_pct", 0)
            nonzero_pct = st.get("nonzero_pct", 0)
            if nonzero_pct >= 20 and null_pct <= 70:
                mark = "★ INTERESSANTE"
            elif null_pct > 70:
                mark = f"MUITOS NULOS ({null_pct}% vazio)"
            else:
                mark = "zero/irrelevante"
            col_lines.append(
                f'"{col}" [número — {mark}] sum={st["sum"]}, avg={st["avg"]}, nonzero={nonzero_pct}%, nulos={null_pct}%'
            )
        elif t == "data":
            col_lines.append(f'"{col}" [data] {st.get("distinct", "?")} valores distintos')
        else:
            top = ", ".join(repr(v) for v in st.get("top3", []))
            col_lines.append(f'"{col}" [texto] {st.get("distinct", "?")} categorias — top: {top}')

    schema_str = (
        f"Dataset: {ds.name} ({total_rows} linhas)\n"
        f"Colunas:\n" + "\n".join(f"  {l}" for l in col_lines) +
        f"\nAmostra ({len(sample)} linhas):\n{json.dumps(sample, ensure_ascii=False, default=str)}"
    )

    system_prompt = (
        "Você é um analista de dados especialista em BI para empresas brasileiras. "
        "Dado o schema e amostra de um dataset, responda a pergunta do usuário retornando SOMENTE um JSON válido "
        "(sem markdown, sem texto fora do JSON) com os campos:\n"
        '{"label_col":"coluna para agrupar (deve ser do tipo texto/categoria)","value_col":"coluna numérica a agregar",'
        '"agg":"sum|count|avg|max|min","filter_col":null,"filter_val":null,'
        '"answer":"resposta em português claro explicando o que o gráfico vai mostrar"}\n'
        "Regras gerais:\n"
        "- Use apenas nomes de colunas exatamente como aparecem no dataset\n"
        "- Prefira colunas do tipo 'número' para value_col e 'texto' para label_col\n"
        "- Para perguntas sobre evolução no tempo, use label_col com coluna de data\n"
        "- Para perguntas de ranking/top, use agg=sum e label_col com categoria\n"
        "- Para médias, use agg=avg\n\n"
        "Conhecimento de domínio (use como referência para interpretar os dados, não como regra fixa):\n"
        "Fiscal brasileiro — NFS-e:\n"
        "- Valor Total Recebido = faturamento bruto; Base de Cálculo = base tributável do ISS\n"
        "- ISS devido = imposto sobre serviço; pode ser zero em empresas do Simples Nacional (imposto unificado no DAS)\n"
        "- PIS, COFINS, CSLL, IR, INSS: zerados no Simples Nacional, com valores no Lucro Presumido/Real — verifique na amostra\n"
        "- Alíquota: taxa percentual, somar não faz sentido; avg faz sentido se o usuário perguntar sobre alíquota média\n"
        "- Para faturamento: sum Valor Total Recebido; para volume: count; para ticket médio: avg Valor Total Recebido\n"
        "- Colunas de data são ideais para análise de tendência temporal"
    )

    client = ant.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model=_AI_MODEL,
        max_tokens=512,
        system=system_prompt,
        messages=[{"role": "user", "content": f"{schema_str}\n\nPergunta: {data.question}"}],
    )

    # Salvar log de uso (não bloqueia em caso de erro)
    try:
        usage = msg.usage
        log = AIUsageLog(
            tenant_id=current_user.tenant_id,
            user_id=current_user.id,
            dataset_id=data.dataset_id,
            model=_AI_MODEL,
            tokens_input=usage.input_tokens,
            tokens_output=usage.output_tokens,
            question=data.question[:500] if data.question else None,
        )
        db.add(log)
        await db.commit()
    except Exception:
        pass

    text = msg.content[0].text.strip()
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if not m:
            raise HTTPException(status_code=500, detail="IA não retornou JSON válido. Tente reformular a pergunta.")
        parsed = json.loads(m.group())

    label_col = parsed.get("label_col")
    value_col = parsed.get("value_col")
    agg = parsed.get("agg", "sum")
    filter_col = parsed.get("filter_col")
    filter_val = parsed.get("filter_val")
    answer = parsed.get("answer", "")

    # ── Normalizar nomes de colunas (case-insensitive + strip) ────────────────
    def _find_col(name: str | None, cols: list[str]) -> str | None:
        if not name:
            return None
        if name in cols:
            return name
        name_n = name.lower().strip()
        for c in cols:
            if c.lower().strip() == name_n:
                return c
        return name  # retorna original — svc.query vai dar 422 com mensagem útil

    label_col = _find_col(label_col, columns)
    value_col = _find_col(value_col, columns)
    filter_col = _find_col(filter_col, columns)

    query_data = svc.query(ds, label_col, value_col, agg, filter_col, filter_val)

    chart_suggestion = _suggest_chart(query_data, data.question)

    return {
        "data": query_data,
        "answer": answer,
        "query": {"label_col": label_col, "value_col": value_col, "agg": agg, "filter_col": filter_col, "filter_val": filter_val},
        **chart_suggestion,
    }


# ---------------------------------------------------------------------------
# Domain detection for BI Advisor
# ---------------------------------------------------------------------------

DOMAIN_TEMPLATES = {
    "fiscal_nfse": {
        "name": "NFS-e / Fiscal",
        "signals": ["nfs", "nfs-e", "nota fiscal", "receita_liquida", "aliquota", "iss", "tomador", "prestador", "rps", "guia", "quitacao"],
        "insights_focus": "faturamento, notas canceladas, concentração de clientes, sazonalidade mensal",
        "kpi_hints": "faturamento total (sum RECEITA_LIQUIDA notas válidas), quantidade de notas, ticket médio, taxa de cancelamento",
        "chart_hints": "evolução mensal (line por MÊS), ranking clientes (bar_h por CLIENTE), mix tipo operação (pie)",
        "optional_improvements": [
            {"col": "CUSTO_OPERACIONAL", "impact": "Calcular margem real (receita - custo)"},
            {"col": "DATA_PAGAMENTO", "impact": "Análise de prazo médio de recebimento (DSO)"},
            {"col": "SEGMENTO_CLIENTE", "impact": "Análise de concentração e LTV por segmento"},
            {"col": "FORMA_PAGAMENTO", "impact": "Mix de recebimento (boleto, PIX, cartão)"},
            {"col": "VENDEDOR", "impact": "Ranking de desempenho por vendedor"},
        ],
    },
    "vendas": {
        "name": "Vendas / Comercial",
        "signals": ["venda", "pedido", "receita", "ticket", "cliente", "produto", "canal", "vendedor", "comissao", "meta"],
        "insights_focus": "receita total, ticket médio, conversão, ranking de produtos e clientes",
        "kpi_hints": "receita total, número de pedidos, ticket médio, clientes ativos",
        "chart_hints": "evolução de vendas (line), ranking produtos (bar_h), mix canal (pie)",
        "optional_improvements": [
            {"col": "CUSTO_PRODUTO", "impact": "Calcular margem bruta por produto"},
            {"col": "DATA_ENTREGA", "impact": "Análise de prazo de entrega e SLA"},
            {"col": "REGIAO", "impact": "Mapa e análise regional de vendas"},
            {"col": "CATEGORIA", "impact": "Análise de mix por categoria"},
        ],
    },
    "marketing": {
        "name": "Marketing / Campanhas",
        "signals": ["campanha", "impressao", "clique", "ctr", "cpl", "cac", "lead", "roas", "investimento", "conversao", "midia"],
        "insights_focus": "ROI de campanhas, custo por lead, eficiência por canal",
        "kpi_hints": "investimento total, leads gerados, CPL, ROAS",
        "chart_hints": "evolução temporal (line), performance por canal (bar_h), mix investimento (pie)",
        "optional_improvements": [
            {"col": "RECEITA_GERADA", "impact": "Calcular ROAS real por campanha"},
            {"col": "TAXA_CONVERSAO", "impact": "Funil de conversão completo"},
        ],
    },
    "rh": {
        "name": "RH / Pessoas",
        "signals": ["colaborador", "funcionario", "salario", "cargo", "departamento", "admissao", "demissao", "folha", "headcount"],
        "insights_focus": "headcount, custo de folha, rotatividade, distribuição por departamento",
        "kpi_hints": "headcount total, custo folha, turnover, salário médio",
        "chart_hints": "headcount por departamento (bar), evolução folha (line), distribuição cargos (pie)",
        "optional_improvements": [
            {"col": "AVALIACAO_DESEMPENHO", "impact": "Correlação entre desempenho e remuneração"},
            {"col": "MOTIVO_SAIDA", "impact": "Análise de causas de turnover"},
        ],
    },
    "financeiro": {
        "name": "Financeiro / DRE",
        "signals": ["receita", "despesa", "custo", "lucro", "margem", "ebitda", "caixa", "dre", "balanco", "resultado"],
        "insights_focus": "resultado líquido, margens, evolução de receita vs despesa",
        "kpi_hints": "receita total, despesas, lucro líquido, margem",
        "chart_hints": "evolução receita vs despesa (line), composição despesas (pie), resultado mensal (bar)",
        "optional_improvements": [
            {"col": "CENTRO_CUSTO", "impact": "Análise de custos por centro de custo"},
            {"col": "CATEGORIA_DESPESA", "impact": "Detalhamento de onde o dinheiro vai"},
        ],
    },
    "ecommerce": {
        "name": "E-commerce",
        "signals": ["pedido", "sku", "produto", "categoria", "frete", "devolucao", "avaliacao", "estoque", "carrinho"],
        "insights_focus": "GMV, taxa de conversão, produtos mais vendidos, devolução",
        "kpi_hints": "GMV total, pedidos, ticket médio, taxa devolução",
        "chart_hints": "evolução GMV (line), top produtos (bar_h), categorias (pie)",
        "optional_improvements": [
            {"col": "MARGEM_PRODUTO", "impact": "Identificar produtos mais rentáveis"},
            {"col": "ORIGEM_TRAFEGO", "impact": "Atribuição de receita por canal de aquisição"},
        ],
    },
}


def _detect_domain(columns: list[str], rows: list[dict]) -> tuple[str, dict]:
    """Detecta o domínio do dataset com base nos nomes das colunas.
    Retorna (domain_key, template_dict).
    """
    cols_lower = " ".join(c.lower() for c in columns)
    best_domain = "generic"
    best_score = 0

    for domain_key, template in DOMAIN_TEMPLATES.items():
        score = sum(1 for signal in template["signals"] if signal in cols_lower)
        if score > best_score:
            best_score = score
            best_domain = domain_key

    if best_score == 0:
        return "generic", {
            "name": "Genérico",
            "insights_focus": "principais métricas e tendências",
            "kpi_hints": "totais e médias das colunas numéricas",
            "chart_hints": "evolução temporal e rankings dimensionais",
            "optional_improvements": [],
        }

    return best_domain, DOMAIN_TEMPLATES[best_domain]


class GenerateDashboardRequest(BaseModel):
    dataset_id: uuid.UUID
    objetivo: str | None = None


@router.post("/generate-dashboard", summary="Gera blocos de dashboard automaticamente com IA")
async def generate_dashboard_endpoint(
    data: GenerateDashboardRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    effective_tenant_id: uuid.UUID = Depends(get_effective_tenant_id),
):
    import json
    import os
    import re
    import uuid as _uuid

    import anthropic as ant

    from .ai_usage_models import AIUsageLog

    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    plan = tenant.plan if tenant else "free"
    check_feature_allowed("ai", plan)
    await check_ai_query_limit(db, current_user.tenant_id, plan)

    # ── Verificação de segurança no objetivo informado ────────────────────────
    if data.objetivo:
        from .security import check_question_safety
        safety = await check_question_safety(
            question=data.objetivo,
            db=db,
            tenant_id=current_user.tenant_id,
            user_id=current_user.id,
            endpoint="generate_dashboard",
        )
        if not safety["safe"]:
            raise HTTPException(status_code=403, detail={
                "code": "safety_violation",
                "message": safety["message"],
                "incident_count": safety.get("incident_count", 1),
            })

    svc = DatasetService(db)
    ds = await svc.get(data.dataset_id, effective_tenant_id)
    if not ds and effective_tenant_id != current_user.tenant_id:
        ds = await svc.get(data.dataset_id, current_user.tenant_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key or api_key == "your_key_here":
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY não configurada.")

    columns = ds.columns or []
    rows = ds.rows or []
    total_rows = len(rows)

    # ── Análise estatística completa de TODAS as linhas ──────────────────────
    col_stats: dict[str, dict] = {}
    for col in columns:
        vals = [r.get(col) for r in rows if r.get(col) is not None and str(r.get(col)).strip().lower() not in (
                "", "null", "none", "n/a", "na", "n.a.", "-", "--", "nan",
                "#n/a", "#na", "#div/0!", "#ref!", "#value!", "#num!", "#name?", "#null!",
            )]
        null_count = total_rows - len(vals)

        if not vals:
            col_stats[col] = {"type": "texto", "null_pct": 100}
            continue

        # Tentar numérico
        nums: list[float] = []
        all_numeric = True
        for v in vals:
            try:
                nums.append(float(str(v).replace(",", ".")))
            except (ValueError, TypeError):
                all_numeric = False
                break

        if all_numeric and nums:
            nonzero = sum(1 for n in nums if n != 0)
            col_stats[col] = {
                "type": "numero",
                "sum": round(sum(nums), 2),
                "avg": round(sum(nums) / len(nums), 2),
                "min": round(min(nums), 2),
                "max": round(max(nums), 2),
                "count": len(nums),
                "nonzero_pct": round(nonzero / len(nums) * 100),
                "null_pct": round(null_count / total_rows * 100) if total_rows else 0,
            }
        else:
            sample_val = str(vals[0])
            if re.match(r"\d{4}-\d{2}|\d{2}/\d{4}|\d{2}/\d{2}/\d{4}|\d{4}/\d{2}/\d{2}", sample_val):
                col_stats[col] = {
                    "type": "data",
                    "sample": sample_val,
                    "count": len(vals),
                }
            else:
                str_vals = [str(v) for v in vals]
                unique_vals = list(dict.fromkeys(str_vals))
                col_stats[col] = {
                    "type": "texto",
                    "unique": len(set(str_vals)),
                    "top3": unique_vals[:3],
                    "count": len(vals),
                }

    # ── Schema para a IA com estatísticas claras ─────────────────────────────
    schema_parts = [
        f"Dataset: {ds.name}",
        f"Total de linhas: {total_rows}",
        "",
        "Colunas (analise cuidadosamente cada uma):",
    ]
    for col, st in col_stats.items():
        if st["type"] == "numero":
            null_pct = st.get("null_pct", 0)
            nonzero_pct = st.get("nonzero_pct", 0)
            # Coluna interessante apenas se: muitos valores preenchidos E maioria não-zero
            # null_pct > 70% = dados insuficientes para gráfico/KPI confiável
            if nonzero_pct >= 20 and null_pct <= 70:
                flag = "★ INTERESSANTE"
            elif null_pct > 70:
                flag = f"— MUITOS NULOS ({null_pct}% vazio)"
            else:
                flag = "— zero/irrelevante"
            schema_parts.append(
                f'  "{col}" [número {flag}] sum={st["sum"]}, avg={st["avg"]}, '
                f'nonzero={nonzero_pct}%, nulos={null_pct}%'
            )
        elif st["type"] == "data":
            schema_parts.append(f'  "{col}" [data] ex: {st["sample"]}')
        else:
            top3 = ", ".join(st.get("top3", []))
            schema_parts.append(
                f'  "{col}" [texto] {st.get("unique", "?")} valores únicos, ex: {top3}'
            )

    sample_rows = rows[:5]
    schema_str = "\n".join(schema_parts)
    schema_str += f"\n\nAmostra (5 linhas):\n{json.dumps(sample_rows, ensure_ascii=False, default=str)}"

    # Detectar coluna de data mais relevante para sugerir ao frontend
    date_cols = [col for col, st in col_stats.items() if st.get("type") == "data"]
    # Preferir colunas com nome relacionado a mês/data/período
    date_priority = ["mes", "mês", "data", "date", "periodo", "período", "dt", "competencia",
                     "emissao", "emissão", "vencimento", "referencia", "referência", "lancamento", "lançamento"]
    suggested_date_col = None
    for priority in date_priority:
        for dc in date_cols:
            if priority in dc.lower():
                suggested_date_col = dc
                break
        if suggested_date_col:
            break
    if not suggested_date_col and date_cols:
        suggested_date_col = date_cols[0]
    # Fallback: checar nome das colunas mesmo que tipo não detectado como data
    if not suggested_date_col:
        for col in columns:
            col_l = col.lower()
            if any(sig in col_l for sig in date_priority):
                suggested_date_col = col
                break

    # ── Detecção de domínio ────────────────────────────────────────────────
    domain_key, domain_tpl = _detect_domain(columns, rows)
    domain_section = (
        f"\nDOMÍNIO DETECTADO: {domain_tpl['name']}\n"
        f"Foco de análise: {domain_tpl['insights_focus']}\n"
        f"KPIs recomendados: {domain_tpl['kpi_hints']}\n"
        f"Gráficos recomendados: {domain_tpl['chart_hints']}\n"
    )

    objetivo_str = f"\n\nObjetivo do usuário: {data.objetivo}" if data.objetivo else ""

    system_prompt = (
        "Você é um Jornalista de Dados com visão estratégica de negócio.\n"
        "Seu trabalho: transformar dados em uma HISTÓRIA VISUAL que o usuário lê de cima para baixo — "
        "como um artigo bem estruturado, não uma coleção de gráficos.\n\n"

        "ANTES DE GERAR, pense:\n"
        "  1. Qual é a PERGUNTA CENTRAL que esses dados respondem?\n"
        "  2. Qual é a HISTÓRIA que os números contam — contexto, tendência, composição, detalhe?\n"
        "  3. Em que ORDEM um leitor descobriria essa história naturalmente, de cima para baixo?\n"
        "  4. Qual é o gráfico HERÓI da história — aquele que sozinho explica o negócio?\n\n"

        "NARRATIVA EM 4 ATOS (os blocos devem estar nessa ordem no JSON):\n\n"
        "  ATO 1 — CONTEXTO: 'Onde estamos agora?'\n"
        "    3–4 KPIs macro — os números mais importantes. O ponto de partida da história.\n"
        "    Ex: Receita Total, Ticket Médio, Volume de Operações, Principal Resultado\n\n"
        "  ATO 2 — TEMPO: 'Como chegamos aqui? Para onde vai?'\n"
        "    1 gráfico temporal herói — linha ou area com a métrica principal ao longo do tempo.\n"
        "    Este é o gráfico mais importante: use w=12 para ele ocupar toda a largura e ter IMPACTO.\n"
        "    Revela tendência, sazonalidade, picos, quedas — a 'espinha dorsal' da história.\n\n"
        "  ATO 3 — COMPOSIÇÃO: 'Quem ou o quê explica esse número?'\n"
        "    Decomposição por dimensão: cliente, produto, região, categoria, canal, tipo.\n"
        "    Use bubble para impacto visual, bar_h para ranking longo, pie para proporção simples.\n"
        "    Responde: 'De onde vem esse resultado? Quem são os protagonistas?'\n\n"
        "  ATO 4 — DETALHE: 'Onde está a oportunidade ou o problema?'\n"
        "    Granularidade fina: segundas métricas, cruzamentos, outliers, distribuições.\n"
        "    O que o CEO perguntaria depois de entender o contexto, a tendência e a composição.\n\n"

        "RETORNE SOMENTE um JSON array válido (sem markdown, sem texto fora do JSON).\n"
        "Os blocos devem estar na ORDEM DA NARRATIVA — o primeiro bloco = início da história.\n\n"
        '[\n'
        '  {"type":"kpi","title":"Receita Total","value_col":"col_valor","agg":"sum","layout":{"w":3,"h":2}},\n'
        '  {"type":"kpi","title":"Ticket Médio","value_col":"col_valor","agg":"avg","layout":{"w":3,"h":2}},\n'
        '  {"type":"line","title":"Evolução Mensal da Receita","label_col":"col_data","value_col":"col_valor","agg":"sum","layout":{"w":12,"h":5}},\n'
        '  {"type":"bubble","title":"Receita por Cliente","label_col":"col_cliente","value_col":"col_valor","agg":"sum","layout":{"w":6,"h":4}},\n'
        '  {"type":"bar_h","title":"Top Produtos por Resultado","label_col":"col_produto","value_col":"col_resultado","agg":"sum","layout":{"w":6,"h":4}},\n'
        '  ...\n'
        ']\n\n'

        "FERRAMENTAS VISUAIS — escolha a que MELHOR SERVE A HISTÓRIA naquele momento:\n"
        "  kpi    → número de impacto. w=3,h=2 para 4 em linha | w=4,h=2 para 3 em linha\n"
        "  line   → série temporal. w=12,h=5 para herói da tela | w=6,h=4 para lado a lado\n"
        "  area   → série temporal acumulada ou com área de destaque\n"
        "  bar    → ranking de 3-8 itens com nomes curtos (w=6,h=4 ou w=4,h=4)\n"
        "  bar_h  → ranking com nomes longos: clientes, produtos, cidades (w=6,h=4)\n"
        "  pie    → proporção de 2–5 categorias com nomes LEGÍVEIS, não códigos (w=4,h=4)\n"
        "  bubble → impacto dimensional: tamanho do círculo = magnitude da métrica (w=6,h=4)\n"
        "  filter → filtro interativo para o usuário explorar (w=4,h=2 ou w=6,h=2)\n\n"

        "DRAMA VISUAL — tamanhos comunicam importância:\n"
        "  w=12 → gráfico herói, ocupa toda a linha, máximo impacto (use no Ato 2)\n"
        "  w=6  → gráfico principal, dois por linha (use no Ato 3 e 4)\n"
        "  w=4  → gráfico de suporte, três por linha\n"
        "  w=3  → KPI, quatro por linha (use no Ato 1)\n"
        "  Altura: h=2 para KPIs/filtros | h=4 para gráficos normais | h=5 para gráficos herói\n\n"

        "REGRAS ABSOLUTAS:\n"
        "1. Use APENAS colunas ★ INTERESSANTE para métricas numéricas\n"
        "2. Ignore colunas '— zero/irrelevante' ou 'MUITOS NULOS'\n"
        "3. Alíquotas/percentuais/taxas: agg=avg — NUNCA sum\n"
        "4. Nomes de colunas EXATAMENTE como no dataset (case-sensitive)\n"
        "5. Títulos são CAPÍTULOS DA HISTÓRIA — informam o que o usuário vai descobrir (máx 36 chars)\n"
        "6. Não repita o mesmo par label_col+value_col em dois blocos\n"
        "7. NUNCA use como dimensão: CNPJ, CPF, ID, código, número de documento, endereço,\n"
        "   CEP, telefone, e-mail, observação, descrição livre — são dados cadastrais, não analíticos\n"
        "8. NUNCA use pie quando top3 mostra letras/códigos isolados (C, P, T, N, S)\n"
        "9. O gráfico do Ato 2 (temporal) deve usar w=12 se houver coluna de data\n\n"

        f"{domain_section}"

        "TESTE DE QUALIDADE — antes de retornar, verifique:\n"
        "  ✓ Os blocos estão na ordem da narrativa (contexto → tempo → composição → detalhe)?\n"
        "  ✓ Um leitor lendo de cima para baixo entende o negócio em 30 segundos?\n"
        "  ✓ Cada bloco responde a pergunta que surge naturalmente após o anterior?\n"
        "  ✓ O gráfico mais importante tem mais espaço (w=12 ou w=6 de destaque)?\n"
        "  ✓ Títulos são informativos como capítulos de artigo, não genéricos ('Gráfico 1')?\n"
        "  ✓ Há pelo menos 1 gráfico temporal com w=12 se o dataset tem coluna de data?"
    )

    client = ant.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model=_AI_MODEL_GENERATE,
        max_tokens=4096,
        system=system_prompt,
        messages=[{"role": "user", "content": f"{schema_str}{objetivo_str}\n\nGere o dashboard JSON:"}],
    )

    # ── Log de uso ────────────────────────────────────────────────────────────
    try:
        usage = msg.usage
        log = AIUsageLog(
            tenant_id=current_user.tenant_id,
            user_id=current_user.id,
            dataset_id=data.dataset_id,
            model=_AI_MODEL_GENERATE,
            tokens_input=usage.input_tokens,
            tokens_output=usage.output_tokens,
            question=f"[generate-dashboard] {data.objetivo or ''}",
        )
        db.add(log)
        await db.commit()
    except Exception:
        pass

    text = msg.content[0].text.strip()
    try:
        blocks = json.loads(text)
    except json.JSONDecodeError:
        m = re.search(r"\[.*\]", text, re.DOTALL)
        if not m:
            raise HTTPException(status_code=500, detail="IA não retornou JSON válido.")
        blocks = json.loads(m.group())

    if not isinstance(blocks, list) or not blocks:
        raise HTTPException(status_code=500, detail="IA retornou formato inválido.")

    # ── Defaults de layout e id por tipo ─────────────────────────────────────
    _layout_defaults = {
        "kpi":    {"w": 3, "h": 2},
        "line":   {"w": 6, "h": 4},
        "area":   {"w": 6, "h": 4},
        "bar":    {"w": 6, "h": 4},
        "bar_h":  {"w": 6, "h": 4},
        "pie":    {"w": 4, "h": 4},
        "bubble": {"w": 6, "h": 4},
    }
    for b in blocks:
        b["dataset_id"] = str(data.dataset_id)
        if not b.get("id"):
            b["id"] = str(_uuid.uuid4())
        btype = b.get("type", "bar")
        default_size = _layout_defaults.get(btype, {"w": 6, "h": 4})
        if not isinstance(b.get("layout"), dict):
            b["layout"] = default_size.copy()
        else:
            b["layout"].setdefault("w", default_size["w"])
            b["layout"].setdefault("h", default_size["h"])

    # ── Posicionamento narrativo dos filtros ──────────────────────────────────
    # Filtro de data: SEMPRE no topo, largura total (w=12) — ancora a narrativa
    # Filtros categoriais: no MEIO da narrativa, após KPIs + gráfico herói temporal,
    #   próximos aos gráficos de composição que eles filtram

    # Remove filtros que a IA possa ter gerado (controlamos o posicionamento)
    blocks = [b for b in blocks if b.get("type") != "filter"]

    # Filtro de data — topo, w=12
    date_filter = []
    if suggested_date_col:
        date_filter.append({
            "id": str(_uuid.uuid4()),
            "type": "filter",
            "config": {"date_mode": True},
            "dataset_id": str(data.dataset_id),
            "filter_col": suggested_date_col,
            "filter_label": suggested_date_col,
            "layout": {"w": 12, "h": 2},
        })

    # Colunas que são identificadores técnicos — NÃO devem virar filtros
    _ID_SIGNALS = ["cnpj", "cpf", "id", "codigo", "código", "chave", "numero", "número",
                   "nfe", "nfse", "rps", "protocolo", "inscricao", "inscrição", "cep",
                   "telefone", "email", "e-mail", "url", "hash", "uuid", "key",
                   "endereco", "endereço", "logradouro", "bairro", "rua", "avenida",
                   "complemento", "descricao", "descrição", "observacao", "observação", "obs"]

    def _is_identifier_col(col_name: str) -> bool:
        cl = col_name.lower().replace(" ", "_").replace("-", "_")
        return any(sig in cl for sig in _ID_SIGNALS)

    def _values_are_codes(top3: list) -> bool:
        for v in top3:
            v = str(v).strip()
            if len(v) <= 2:
                return True
            if v.isupper() and len(v) <= 6 and " " not in v:
                return True
        return False

    # Filtros categoriais — até 2, colocados no MEIO da narrativa
    cat_filters = []
    cat_filters_added = 0
    for col, st in col_stats.items():
        if cat_filters_added >= 2:
            break
        if col == suggested_date_col:
            continue
        if _is_identifier_col(col):
            continue
        if st.get("type") == "texto":
            unique = st.get("unique", 0)
            top3 = st.get("top3", [])
            if 2 <= unique <= 20 and not _values_are_codes(top3):
                cat_filters.append({
                    "id": str(_uuid.uuid4()),
                    "type": "filter",
                    "dataset_id": str(data.dataset_id),
                    "filter_col": col,
                    "filter_label": col,
                    "layout": {"w": 6, "h": 2},
                })
                cat_filters_added += 1

    # Ponto de inserção dos filtros categoriais:
    # → após todos os KPIs e o primeiro gráfico temporal (line/area)
    # → antes dos gráficos de composição (bubble, bar_h, pie)
    insert_after = 0
    found_temporal = False
    for i, b in enumerate(blocks):
        if b.get("type") == "kpi":
            insert_after = i + 1
        elif b.get("type") in ("line", "area") and not found_temporal:
            insert_after = i + 1
            found_temporal = True
        elif found_temporal:
            break  # primeiro bloco não-temporal após o herói = ponto certo

    # Ordem narrativa final:
    # filtro_data → KPIs → gráfico_herói → filtros_categoria → composição → detalhe
    blocks = date_filter + blocks[:insert_after] + cat_filters + blocks[insert_after:]

    return {
        "blocks": blocks,
        "suggested_date_col": suggested_date_col,
        "domain": domain_key,
        "domain_name": domain_tpl["name"],
    }


@router.get("/ai-usage", summary="Retorna cota mensal de IA do tenant")
async def get_ai_usage(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    import uuid as _uuid
    from datetime import timezone
    from sqlalchemy import func as _func, select as _sel
    from app.modules.billing.plan_limits import PLANS
    from .ai_usage_models import AIUsageLog

    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    plan = tenant.plan if tenant else "free"
    limits = PLANS.get(plan, PLANS["free"])
    max_queries = limits.max_ai_queries_monthly

    now = __import__("datetime").datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    # primeiro dia do próximo mês
    if now.month == 12:
        next_month = now.replace(year=now.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        next_month = now.replace(month=now.month + 1, day=1, hour=0, minute=0, second=0, microsecond=0)

    used = 0
    if limits.allow_ai:
        used = await db.scalar(
            _sel(_func.count()).select_from(AIUsageLog).where(
                AIUsageLog.tenant_id == current_user.tenant_id,
                AIUsageLog.created_at >= month_start,
            )
        ) or 0

    return {
        "used": used,
        "limit": max_queries,       # -1 = ilimitado
        "remaining": -1 if max_queries == -1 else max(0, max_queries - used),
        "reset_at": next_month.isoformat(),
        "plan": plan,
    }


# ---------------------------------------------------------------------------
# Alertas por threshold
# ---------------------------------------------------------------------------

from datetime import datetime, timezone as _tz

from sqlalchemy import select as _select

from .alert_models import ReportAlert


class AlertCreate(BaseModel):
    name: str
    dataset_id: uuid.UUID
    value_col: str
    agg: str = "sum"
    operator: str  # gt | lt | gte | lte | eq
    threshold: float
    filter_col: str | None = None
    filter_val: str | None = None
    notify_email: str | None = None
    notify_slack_url: str | None = None
    notify_sms: str | None = None


class AlertResponse(BaseModel):
    id: uuid.UUID
    name: str
    dataset_id: uuid.UUID
    value_col: str
    agg: str
    operator: str
    threshold: float
    filter_col: str | None
    filter_val: str | None
    is_active: bool
    last_value: float | None
    last_status: str | None
    checked_at: str | None
    created_at: str
    notify_email: str | None = None
    notify_slack_url: str | None = None
    notify_sms: str | None = None

    model_config = {"from_attributes": True}


async def _send_alert_notifications(alert: ReportAlert, value: float) -> None:
    """Envia notificações por email e/ou Slack quando alerta é disparado."""
    msg = f"Alerta '{alert.name}' disparado! Valor atual: {value:.2f} (threshold: {alert.threshold})"

    if alert.notify_email:
        try:
            from app.core.email import send_generic_email
            for email in alert.notify_email.split(","):
                email = email.strip()
                if email:
                    await send_generic_email(email, f"Alerta Jarbis: {alert.name}", msg)
        except Exception as e:
            print(f"[ALERT EMAIL] Falha: {e}")

    if alert.notify_slack_url:
        try:
            import httpx
            async with httpx.AsyncClient(timeout=10) as client:
                await client.post(alert.notify_slack_url, json={"text": msg})
        except Exception as e:
            print(f"[ALERT SLACK] Falha: {e}")

    # SMS via Twilio
    if alert.notify_sms:
        try:
            import os
            from twilio.rest import Client as TwilioClient
            from app.config import settings as _settings
            twilio_sid = os.environ.get("TWILIO_ACCOUNT_SID")
            twilio_token = os.environ.get("TWILIO_AUTH_TOKEN")
            twilio_from = os.environ.get("TWILIO_FROM_NUMBER")
            if twilio_sid and twilio_token and twilio_from:
                tw = TwilioClient(twilio_sid, twilio_token)
                sms_body = f"🚨 Alerta Jarbis: {alert.name}\nCondição atingida no dashboard.\nAcesse: {_settings.frontend_url}"
                for phone in (alert.notify_sms or "").split(","):
                    phone = phone.strip()
                    if phone:
                        tw.messages.create(body=sms_body, from_=twilio_from, to=phone)
        except Exception as e:
            print(f"[SMS] Erro: {e}")


def _eval_alert(alert: ReportAlert, current_value: float) -> str:
    op = alert.operator
    t = alert.threshold
    if op == "gt":   return "triggered" if current_value > t else "ok"
    if op == "lt":   return "triggered" if current_value < t else "ok"
    if op == "gte":  return "triggered" if current_value >= t else "ok"
    if op == "lte":  return "triggered" if current_value <= t else "ok"
    if op == "eq":   return "triggered" if current_value == t else "ok"
    return "ok"


@router.get("/alerts", response_model=list[AlertResponse], summary="Lista alertas do tenant")
async def list_alerts(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    result = await db.scalars(
        _select(ReportAlert)
        .where(ReportAlert.tenant_id == current_user.tenant_id)
        .order_by(ReportAlert.created_at.desc())
    )
    alerts = result.all()
    return [AlertResponse(
        id=a.id, name=a.name, dataset_id=a.dataset_id, value_col=a.value_col,
        agg=a.agg, operator=a.operator, threshold=a.threshold,
        filter_col=a.filter_col, filter_val=a.filter_val, is_active=a.is_active,
        last_value=a.last_value, last_status=a.last_status,
        checked_at=a.checked_at.isoformat() if a.checked_at else None,
        created_at=a.created_at.isoformat(),
    ) for a in alerts]


@router.post("/alerts", response_model=AlertResponse, status_code=201, summary="Cria alerta")
async def create_alert(
    data: AlertCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    await check_alert_limit(db, current_user.tenant_id, tenant.plan if tenant else "free", tenant.addon_packs if tenant else 0)
    valid_ops = {"gt", "lt", "gte", "lte", "eq"}
    if data.operator not in valid_ops:
        raise HTTPException(status_code=422, detail=f"Operador inválido. Use: {', '.join(valid_ops)}")

    # Validate dataset belongs to tenant
    ds_svc = DatasetService(db)
    ds = await ds_svc.get(data.dataset_id, current_user.tenant_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")

    alert = ReportAlert(
        tenant_id=current_user.tenant_id,
        dataset_id=data.dataset_id,
        name=data.name,
        value_col=data.value_col,
        agg=data.agg,
        operator=data.operator,
        threshold=data.threshold,
        filter_col=data.filter_col,
        filter_val=data.filter_val,
        notify_email=data.notify_email,
        notify_slack_url=data.notify_slack_url,
        notify_sms=data.notify_sms,
    )
    db.add(alert)
    await db.flush()
    return AlertResponse(
        id=alert.id, name=alert.name, dataset_id=alert.dataset_id, value_col=alert.value_col,
        agg=alert.agg, operator=alert.operator, threshold=alert.threshold,
        filter_col=alert.filter_col, filter_val=alert.filter_val, is_active=alert.is_active,
        last_value=None, last_status=None, checked_at=None,
        created_at=alert.created_at.isoformat(),
    )


@router.post("/alerts/{alert_id}/check", response_model=AlertResponse, summary="Avalia alerta agora")
async def check_alert(
    alert_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    alert = await db.scalar(
        _select(ReportAlert).where(
            ReportAlert.id == alert_id,
            ReportAlert.tenant_id == current_user.tenant_id,
        )
    )
    if not alert:
        raise HTTPException(status_code=404, detail="Alerta não encontrado")

    ds_svc = DatasetService(db)
    ds = await ds_svc.get(alert.dataset_id, current_user.tenant_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")

    rows = ds_svc.query(ds, alert.value_col, alert.value_col, alert.agg, alert.filter_col, alert.filter_val)
    total = sum(r.get("value", 0) or 0 for r in rows)

    prev_status = alert.last_status
    new_status = _eval_alert(alert, total)
    alert.last_value = total
    alert.last_status = new_status
    alert.checked_at = datetime.now(_tz.utc)
    await db.commit()
    await db.refresh(alert)

    # Dispara notificações se passou para triggered
    if new_status == "triggered" and prev_status != "triggered":
        await _send_alert_notifications(alert, total)

    return AlertResponse(
        id=alert.id, name=alert.name, dataset_id=alert.dataset_id, value_col=alert.value_col,
        agg=alert.agg, operator=alert.operator, threshold=alert.threshold,
        filter_col=alert.filter_col, filter_val=alert.filter_val, is_active=alert.is_active,
        last_value=alert.last_value, last_status=alert.last_status,
        checked_at=alert.checked_at.isoformat() if alert.checked_at else None,
        created_at=alert.created_at.isoformat(),
        notify_email=alert.notify_email, notify_slack_url=alert.notify_slack_url,
    )


@router.patch("/alerts/{alert_id}", response_model=AlertResponse, summary="Ativa/desativa alerta")
async def toggle_alert(
    alert_id: uuid.UUID,
    is_active: bool,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    alert = await db.scalar(
        _select(ReportAlert).where(
            ReportAlert.id == alert_id,
            ReportAlert.tenant_id == current_user.tenant_id,
        )
    )
    if not alert:
        raise HTTPException(status_code=404, detail="Alerta não encontrado")
    alert.is_active = is_active
    await db.commit()
    await db.refresh(alert)
    return AlertResponse(
        id=alert.id, name=alert.name, dataset_id=alert.dataset_id, value_col=alert.value_col,
        agg=alert.agg, operator=alert.operator, threshold=alert.threshold,
        filter_col=alert.filter_col, filter_val=alert.filter_val, is_active=alert.is_active,
        last_value=alert.last_value, last_status=alert.last_status,
        checked_at=alert.checked_at.isoformat() if alert.checked_at else None,
        created_at=alert.created_at.isoformat(),
    )


@router.delete("/alerts/{alert_id}", status_code=204, summary="Remove alerta")
async def delete_alert(
    alert_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    alert = await db.scalar(
        _select(ReportAlert).where(
            ReportAlert.id == alert_id,
            ReportAlert.tenant_id == current_user.tenant_id,
        )
    )
    if not alert:
        raise HTTPException(status_code=404, detail="Alerta não encontrado")
    await db.delete(alert)
    await db.commit()


# ---------------------------------------------------------------------------
# Collections (pastas de dashboards) — ANTES de /{report_id} para não ser capturado
# ---------------------------------------------------------------------------


class CollectionCreate(BaseModel):
    name: str
    description: str | None = None
    color: str = "#7c3aed"


class CollectionResponse(BaseModel):
    id: uuid.UUID
    name: str
    description: str | None
    color: str | None
    is_pinned: bool
    created_at: datetime
    report_count: int = 0

    class Config:
        from_attributes = True


@router.get("/collections", response_model=list[CollectionResponse])
async def list_collections(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    from .collection_models import Collection, report_collection
    from sqlalchemy import func
    result = await db.execute(
        select(Collection, func.count(report_collection.c.report_id).label('cnt'))
        .outerjoin(report_collection, Collection.id == report_collection.c.collection_id)
        .where(Collection.tenant_id == current_user.tenant_id)
        .group_by(Collection.id)
        .order_by(Collection.is_pinned.desc(), Collection.created_at.desc())
    )
    rows = result.all()
    return [CollectionResponse(
        id=c.id, name=c.name, description=c.description, color=c.color,
        is_pinned=c.is_pinned, created_at=c.created_at, report_count=cnt
    ) for c, cnt in rows]


@router.post("/collections", response_model=CollectionResponse, status_code=201)
async def create_collection(
    body: CollectionCreate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    from .collection_models import Collection
    col = Collection(tenant_id=current_user.tenant_id, **body.model_dump())
    db.add(col)
    await db.commit()
    await db.refresh(col)
    return CollectionResponse(
        **{k: getattr(col, k) for k in ['id', 'name', 'description', 'color', 'is_pinned', 'created_at']},
        report_count=0,
    )


@router.patch("/collections/{collection_id}")
async def update_collection(
    collection_id: uuid.UUID,
    body: CollectionCreate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    from .collection_models import Collection
    col = await db.scalar(
        select(Collection).where(Collection.id == collection_id, Collection.tenant_id == current_user.tenant_id)
    )
    if not col:
        raise HTTPException(status_code=404, detail="Coleção não encontrada")
    col.name = body.name
    if body.color:
        col.color = body.color
    if body.description is not None:
        col.description = body.description
    await db.commit()
    await db.refresh(col)
    return CollectionResponse(
        **{k: getattr(col, k) for k in ['id', 'name', 'description', 'color', 'is_pinned', 'created_at']},
        report_count=0,
    )


@router.delete("/collections/{collection_id}")
async def delete_collection(
    collection_id: uuid.UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    from .collection_models import Collection
    col = await db.scalar(
        select(Collection).where(Collection.id == collection_id, Collection.tenant_id == current_user.tenant_id)
    )
    if not col:
        raise HTTPException(status_code=404, detail="Coleção não encontrada")
    await db.delete(col)
    await db.commit()
    return {"ok": True}


@router.post("/collections/{collection_id}/reports/{report_id}")
async def add_report_to_collection(
    collection_id: uuid.UUID,
    report_id: uuid.UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    from .collection_models import report_collection
    from sqlalchemy.dialects.postgresql import insert as pg_insert
    await db.execute(
        pg_insert(report_collection)
        .values(report_id=report_id, collection_id=collection_id)
        .on_conflict_do_nothing()
    )
    await db.commit()
    return {"ok": True}


@router.delete("/collections/{collection_id}/reports/{report_id}")
async def remove_report_from_collection(
    collection_id: uuid.UUID,
    report_id: uuid.UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    from .collection_models import report_collection
    await db.execute(
        report_collection.delete().where(
            report_collection.c.collection_id == collection_id,
            report_collection.c.report_id == report_id,
        )
    )
    await db.commit()
    return {"ok": True}


@router.get("/collections/{collection_id}/reports")
async def list_collection_reports(
    collection_id: uuid.UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    from .collection_models import Collection, report_collection
    from .models import Report
    col = await db.scalar(
        select(Collection).where(Collection.id == collection_id, Collection.tenant_id == current_user.tenant_id)
    )
    if not col:
        raise HTTPException(status_code=404, detail="Coleção não encontrada")
    result = await db.execute(
        select(Report)
        .join(report_collection, Report.id == report_collection.c.report_id)
        .where(
            report_collection.c.collection_id == collection_id,
            Report.tenant_id == current_user.tenant_id,
        )
        .order_by(Report.updated_at.desc())
    )
    reports = result.scalars().all()
    return [{"id": str(r.id), "title": r.title, "cover_image": getattr(r, 'cover_image', None)} for r in reports]


# ---------------------------------------------------------------------------
# Clone — deve ficar ANTES de /{report_id} para não ser capturado
# ---------------------------------------------------------------------------


@router.post(
    "/{report_id}/clone",
    response_model=ReportResponse,
    status_code=201,
    summary="Clona um relatório existente",
)
async def clone_report(
    report_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    await check_dashboard_limit(db, current_user.tenant_id, tenant.plan if tenant else "free", tenant.addon_packs if tenant else 0)
    service = ReportService(db)
    return await service.clone(report_id, current_user.tenant_id)


# ---------------------------------------------------------------------------
# Phase 8 — Notificações in-app
# ---------------------------------------------------------------------------

from app.modules.admin.models import UserNotification, TenantWebhook, ApiKey  # noqa: E402


@router.get("/notifications", summary="Listar notificações do usuário")
async def list_notifications(
    unread_only: bool = Query(default=False),
    limit: int = Query(default=20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    q = select(UserNotification).where(
        UserNotification.user_id == current_user.id,
        UserNotification.tenant_id == current_user.tenant_id,
    )
    if unread_only:
        q = q.where(UserNotification.read_at.is_(None))
    q = q.order_by(UserNotification.created_at.desc()).limit(limit)
    rows = await db.scalars(q)
    items = rows.all()
    return {
        "items": [
            {
                "id": str(n.id),
                "type": n.type,
                "title": n.title,
                "body": n.body,
                "link": n.link,
                "read": n.read_at is not None,
                "created_at": n.created_at.isoformat(),
            }
            for n in items
        ],
        "unread_count": sum(1 for n in items if n.read_at is None),
    }


@router.post("/notifications/read-all", summary="Marcar todas as notificações como lidas")
async def mark_all_notifications_read(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    await db.execute(
        update(UserNotification)
        .where(
            UserNotification.user_id == current_user.id,
            UserNotification.read_at.is_(None),
        )
        .values(read_at=datetime.now(timezone.utc))
    )
    await db.commit()
    return {"ok": True}


@router.post("/notifications/{notification_id}/read", summary="Marcar notificação como lida")
async def mark_notification_read(
    notification_id: uuid.UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    n = await db.scalar(
        select(UserNotification).where(
            UserNotification.id == notification_id,
            UserNotification.user_id == current_user.id,
        )
    )
    if not n:
        raise HTTPException(status_code=404, detail="Notificação não encontrada.")
    if not n.read_at:
        n.read_at = datetime.now(timezone.utc)
        await db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Phase 8 — Webhooks outbound
# ---------------------------------------------------------------------------


class WebhookCreateInput(BaseModel):
    url: str
    events: str = "alert.triggered"


class WebhookUpdateInput(BaseModel):
    url: str | None = None
    events: str | None = None
    is_active: bool | None = None


@router.get("/tenant/webhooks", summary="Listar webhooks do tenant")
async def list_webhooks(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    rows = await db.scalars(
        select(TenantWebhook)
        .where(TenantWebhook.tenant_id == current_user.tenant_id)
        .order_by(TenantWebhook.created_at.desc())
    )
    items = rows.all()
    return {
        "items": [
            {
                "id": str(w.id),
                "url": w.url,
                "events": w.events,
                "is_active": w.is_active,
                "last_triggered_at": w.last_triggered_at.isoformat() if w.last_triggered_at else None,
                "last_status_code": w.last_status_code,
                "created_by": w.created_by,
                "created_at": w.created_at.isoformat(),
            }
            for w in items
        ]
    }


@router.post("/tenant/webhooks", status_code=201, summary="Criar webhook")
async def create_webhook(
    data: WebhookCreateInput,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    if tenant and tenant.plan not in ("ilimitado", "enterprise"):
        raise HTTPException(status_code=403, detail="Webhooks disponíveis apenas no plano Ilimitado.")
    secret = secrets.token_hex(32)
    webhook = TenantWebhook(
        tenant_id=current_user.tenant_id,
        url=data.url,
        events=data.events,
        secret=secret,
        is_active=True,
        created_by=current_user.email,
    )
    db.add(webhook)
    await db.commit()
    await db.refresh(webhook)
    return {"id": str(webhook.id), "secret": secret}


@router.patch("/tenant/webhooks/{webhook_id}", summary="Atualizar webhook")
async def update_webhook(
    webhook_id: uuid.UUID,
    data: WebhookUpdateInput,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    w = await db.scalar(
        select(TenantWebhook).where(
            TenantWebhook.id == webhook_id,
            TenantWebhook.tenant_id == current_user.tenant_id,
        )
    )
    if not w:
        raise HTTPException(status_code=404, detail="Webhook não encontrado.")
    if data.url is not None:
        w.url = data.url
    if data.events is not None:
        w.events = data.events
    if data.is_active is not None:
        w.is_active = data.is_active
    await db.commit()
    return {"ok": True}


@router.delete("/tenant/webhooks/{webhook_id}", summary="Deletar webhook")
async def delete_webhook(
    webhook_id: uuid.UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    w = await db.scalar(
        select(TenantWebhook).where(
            TenantWebhook.id == webhook_id,
            TenantWebhook.tenant_id == current_user.tenant_id,
        )
    )
    if not w:
        raise HTTPException(status_code=404, detail="Webhook não encontrado.")
    await db.delete(w)
    await db.commit()
    return {"ok": True}


@router.post("/tenant/webhooks/{webhook_id}/test", summary="Testar webhook")
async def test_webhook(
    webhook_id: uuid.UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    w = await db.scalar(
        select(TenantWebhook).where(
            TenantWebhook.id == webhook_id,
            TenantWebhook.tenant_id == current_user.tenant_id,
        )
    )
    if not w:
        raise HTTPException(status_code=404, detail="Webhook não encontrado.")
    payload = {
        "event": "webhook.test",
        "tenant_id": str(current_user.tenant_id),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    body = json.dumps(payload).encode()
    sig = hmac.new(w.secret.encode(), body, hashlib.sha256).hexdigest()
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.post(
                w.url,
                content=body,
                headers={"Content-Type": "application/json", "X-Jarbis-Signature": f"sha256={sig}"},
            )
        w.last_triggered_at = datetime.now(timezone.utc)
        w.last_status_code = resp.status_code
        await db.commit()
        return {"ok": True, "status_code": resp.status_code}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ---------------------------------------------------------------------------
# Phase 8 — White-label / Customização
# ---------------------------------------------------------------------------


class CustomizationUpdateInput(BaseModel):
    custom_logo_url: str | None = None
    primary_color: str | None = None


@router.get("/tenant/customization", summary="Configurações de personalização")
async def get_customization(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant não encontrado.")
    return {
        "custom_logo_url": tenant.custom_logo_url,
        "primary_color": tenant.primary_color,
        "plan": tenant.plan,
        "white_label_enabled": tenant.plan in ("ilimitado", "enterprise"),
    }


@router.patch("/tenant/customization", summary="Atualizar personalização (white-label)")
async def update_customization(
    data: CustomizationUpdateInput,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant não encontrado.")
    if tenant.plan not in ("ilimitado", "enterprise"):
        raise HTTPException(status_code=403, detail="White-label disponível apenas no plano Ilimitado.")
    if data.custom_logo_url is not None:
        tenant.custom_logo_url = data.custom_logo_url
    if data.primary_color is not None:
        if data.primary_color and not (len(data.primary_color) == 7 and data.primary_color.startswith("#")):
            raise HTTPException(status_code=400, detail="Cor deve ser no formato #RRGGBB.")
        tenant.primary_color = data.primary_color
    await db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Phase 8 — Templates de dashboard
# ---------------------------------------------------------------------------

DASHBOARD_TEMPLATES = [
    {
        "id": "sales-overview",
        "name": "Visão Geral de Vendas",
        "description": "KPIs de faturamento, pedidos e ticket médio com gráfico de evolução.",
        "category": "vendas",
        "preview_color": "#7c3aed",
        "pages": [{"id": "p1", "title": "Vendas", "blocks": [
            {"id": "b1", "type": "metric", "x": 0, "y": 0, "w": 3, "h": 2, "config": {"title": "Faturamento", "format": "currency"}},
            {"id": "b2", "type": "metric", "x": 3, "y": 0, "w": 3, "h": 2, "config": {"title": "Pedidos", "format": "number"}},
            {"id": "b3", "type": "metric", "x": 6, "y": 0, "w": 3, "h": 2, "config": {"title": "Ticket Médio", "format": "currency"}},
            {"id": "b4", "type": "bar", "x": 0, "y": 2, "w": 12, "h": 4, "config": {"title": "Faturamento por Período", "x_field": "data", "y_field": "valor"}},
        ]}],
    },
    {
        "id": "marketing-analytics",
        "name": "Analytics de Marketing",
        "description": "Funil de conversão, canais de aquisição e performance de campanhas.",
        "category": "marketing",
        "preview_color": "#0891b2",
        "pages": [{"id": "p1", "title": "Marketing", "blocks": [
            {"id": "b1", "type": "pie", "x": 0, "y": 0, "w": 6, "h": 4, "config": {"title": "Canais de Aquisição", "field": "canal"}},
            {"id": "b2", "type": "bar", "x": 6, "y": 0, "w": 6, "h": 4, "config": {"title": "Conversões por Campanha", "x_field": "campanha", "y_field": "conversoes"}},
        ]}],
    },
    {
        "id": "financial-dashboard",
        "name": "Dashboard Financeiro",
        "description": "Receitas, despesas, margem e fluxo de caixa em um só lugar.",
        "category": "financeiro",
        "preview_color": "#059669",
        "pages": [{"id": "p1", "title": "Financeiro", "blocks": [
            {"id": "b1", "type": "metric", "x": 0, "y": 0, "w": 4, "h": 2, "config": {"title": "Receita Bruta", "format": "currency"}},
            {"id": "b2", "type": "metric", "x": 4, "y": 0, "w": 4, "h": 2, "config": {"title": "Despesas", "format": "currency"}},
            {"id": "b3", "type": "metric", "x": 8, "y": 0, "w": 4, "h": 2, "config": {"title": "Margem Líquida", "format": "percent"}},
            {"id": "b4", "type": "line", "x": 0, "y": 2, "w": 12, "h": 4, "config": {"title": "Evolução Financeira", "x_field": "mes", "y_field": "valor"}},
        ]}],
    },
    {
        "id": "operations-kpi",
        "name": "KPIs Operacionais",
        "description": "Indicadores operacionais: SLA, satisfação, volume de atendimentos.",
        "category": "operacoes",
        "preview_color": "#d97706",
        "pages": [{"id": "p1", "title": "Operações", "blocks": [
            {"id": "b1", "type": "metric", "x": 0, "y": 0, "w": 3, "h": 2, "config": {"title": "Atendimentos", "format": "number"}},
            {"id": "b2", "type": "metric", "x": 3, "y": 0, "w": 3, "h": 2, "config": {"title": "SLA Cumprido", "format": "percent"}},
            {"id": "b3", "type": "metric", "x": 6, "y": 0, "w": 3, "h": 2, "config": {"title": "Satisfação (NPS)", "format": "number"}},
            {"id": "b4", "type": "table", "x": 0, "y": 2, "w": 12, "h": 5, "config": {"title": "Detalhamento por Equipe"}},
        ]}],
    },
]


@router.get("/templates", summary="Listar templates de dashboard")
async def list_templates(
    current_user: User = Depends(get_current_active_user),
):
    return {"items": DASHBOARD_TEMPLATES}


@router.post("/from-template/{template_id}", status_code=201, summary="Criar dashboard a partir de template")
async def create_from_template(
    template_id: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    template = next((t for t in DASHBOARD_TEMPLATES if t["id"] == template_id), None)
    if not template:
        raise HTTPException(status_code=404, detail="Template não encontrado.")
    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    await check_dashboard_limit(db, current_user.tenant_id, tenant.plan if tenant else "free", tenant.addon_packs if tenant else 0)
    service = ReportService(db)
    report = await service.create(
        current_user.tenant_id,
        ReportCreate(title=template["name"], description=template["description"]),
    )
    report.pages = template["pages"]
    await db.commit()
    await db.refresh(report)
    return {"id": str(report.id), "title": report.title}


# ---------------------------------------------------------------------------
# API Keys — tenant-facing (self-service)
# ---------------------------------------------------------------------------


class ApiKeyCreate(BaseModel):
    name: str
    scopes: str = "read"
    expires_in_days: int | None = None


@router.get("/api-keys", summary="Lista API Keys do tenant")
async def list_api_keys(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    rows = await db.scalars(
        select(ApiKey)
        .where(ApiKey.tenant_id == current_user.tenant_id)
        .order_by(ApiKey.created_at.desc())
    )
    return [
        {
            "id": str(k.id),
            "name": k.name,
            "key_prefix": k.key_prefix,
            "scopes": k.scopes,
            "is_active": k.is_active,
            "last_used_at": k.last_used_at.isoformat() if k.last_used_at else None,
            "expires_at": k.expires_at.isoformat() if k.expires_at else None,
            "created_at": k.created_at.isoformat() if k.created_at else None,
        }
        for k in rows.all()
    ]


@router.post("/api-keys", summary="Cria nova API Key para o tenant", status_code=201)
async def create_api_key(
    data: ApiKeyCreate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    raw_key = f"jrb_live_{secrets.token_hex(20)}"
    hashed = hashlib.sha256(raw_key.encode()).hexdigest()
    prefix = raw_key[:12]
    expires_at = None
    if data.expires_in_days:
        from datetime import timedelta
        expires_at = datetime.now(timezone.utc) + timedelta(days=data.expires_in_days)
    key = ApiKey(
        tenant_id=current_user.tenant_id,
        name=data.name,
        key_prefix=prefix,
        hashed_key=hashed,
        scopes=data.scopes,
        expires_at=expires_at,
        created_by=current_user.email,
    )
    db.add(key)
    await db.commit()
    await db.refresh(key)
    return {
        "id": str(key.id),
        "name": key.name,
        "key": raw_key,
        "key_prefix": key.key_prefix,
        "scopes": key.scopes,
        "expires_at": key.expires_at.isoformat() if key.expires_at else None,
        "created_at": key.created_at.isoformat() if key.created_at else None,
    }


@router.delete("/api-keys/{key_id}", summary="Revoga (desativa) uma API Key", status_code=204)
async def revoke_api_key(
    key_id: uuid.UUID,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    key = await db.scalar(
        select(ApiKey).where(ApiKey.id == key_id, ApiKey.tenant_id == current_user.tenant_id)
    )
    if not key:
        raise HTTPException(status_code=404, detail="Chave não encontrada.")
    key.is_active = False
    await db.commit()


# ---------------------------------------------------------------------------
# Embed RLS tokens — geração e decodificação de tokens com filtros por usuário
# ---------------------------------------------------------------------------


class EmbedTokenCreate(BaseModel):
    dashboard_id: uuid.UUID
    user_ref: str  # external user identifier
    filters: list[dict] = []  # [{col: str, value: str}]
    expires_in_hours: int = 4


class EmbedTokenResponse(BaseModel):
    token: str
    dashboard_id: uuid.UUID
    expires_at: str
    embed_url: str


@router.post("/embed-tokens", response_model=EmbedTokenResponse)
async def create_embed_token(
    body: EmbedTokenCreate,
    current_user=Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Cria token embed com filtros RLS por usuário final."""
    import jwt as pyjwt
    from datetime import timedelta
    from app.config import settings

    now = datetime.now(timezone.utc)
    expires_at = now + timedelta(hours=body.expires_in_hours)

    payload = {
        "sub": str(body.dashboard_id),
        "tenant_id": str(current_user.tenant_id),
        "user_ref": body.user_ref,
        "filters": body.filters,
        "exp": expires_at,
        "iat": now,
        "type": "embed",
    }

    secret = getattr(settings, "secret_key", "jarbis-secret")
    token = pyjwt.encode(payload, secret, algorithm="HS256")

    return EmbedTokenResponse(
        token=token,
        dashboard_id=body.dashboard_id,
        expires_at=expires_at.isoformat(),
        embed_url=f"https://app.jarbis.cc/embed/{token}",
    )


@router.get("/embed/{token}")
async def decode_embed_token(token: str):
    """Decodifica token embed e retorna dados do dashboard com filtros."""
    import jwt as pyjwt
    from app.config import settings

    try:
        secret = getattr(settings, "secret_key", "jarbis-secret")
        payload = pyjwt.decode(token, secret, algorithms=["HS256"])
        if payload.get("type") != "embed":
            raise HTTPException(status_code=400, detail="Token inválido.")
        return {
            "dashboard_id": payload["sub"],
            "tenant_id": payload["tenant_id"],
            "user_ref": payload.get("user_ref"),
            "filters": payload.get("filters", []),
        }
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado.")
    except Exception:
        raise HTTPException(status_code=400, detail="Token inválido.")


# ---------------------------------------------------------------------------
# N30 — Google Analytics Connector
# ---------------------------------------------------------------------------

from app.modules.reports.connectors.ga_connector import fetch_ga_report


class GADatasetCreate(BaseModel):
    name: str
    property_id: str
    api_key: str
    dimensions: list[str] = ["date", "sessionDefaultChannelGrouping"]
    metrics: list[str] = ["sessions", "activeUsers", "bounceRate"]
    date_range_days: int = 30


@router.post("/datasets/google-analytics", status_code=201)
async def create_ga_dataset(
    body: GADatasetCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_active_user),
):
    """Create a dataset from Google Analytics Data API."""
    try:
        rows = await fetch_ga_report(
            property_id=body.property_id,
            api_secret=body.api_key,
            dimensions=body.dimensions,
            metrics=body.metrics,
            date_range_days=body.date_range_days,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao buscar dados do GA: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="Nenhum dado retornado. Verifique Property ID e API Key.")

    columns = list(rows[0].keys()) if rows else []
    dataset = ReportDataset(
        id=uuid.uuid4(),
        tenant_id=user.tenant_id,
        name=body.name,
        type="api",
        api_url=f"https://analyticsdata.googleapis.com/v1beta/properties/{body.property_id}:runReport",
        rows=rows,
        columns=columns,
        row_count=len(rows),
        sync_mode="replace",
    )
    db.add(dataset)
    await db.commit()
    return {"id": str(dataset.id), "name": dataset.name, "row_count": len(rows), "columns": columns}


# ---------------------------------------------------------------------------
# G2 — Relatórios Agendados
# ---------------------------------------------------------------------------

from .models import Report
from .scheduled_models import ScheduledReport
from .scheduled_service import ScheduledReportService


# ---------------------------------------------------------------------------
# N7 — Version History for dashboards
# ---------------------------------------------------------------------------


class VersionResponse(BaseModel):
    id: uuid.UUID
    report_id: uuid.UUID
    title: str
    label: str | None
    created_at: datetime

    class Config:
        from_attributes = True


@router.get("/{report_id}/versions", response_model=list[VersionResponse])
async def list_versions(
    report_id: uuid.UUID,
    current_user=Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    from .version_models import ReportVersion
    result = await db.execute(
        select(ReportVersion)
        .where(ReportVersion.report_id == report_id, ReportVersion.tenant_id == current_user.tenant_id)
        .order_by(ReportVersion.created_at.desc())
        .limit(20)
    )
    return result.scalars().all()


@router.post("/{report_id}/versions")
async def save_version(
    report_id: uuid.UUID,
    label: str | None = Query(None),
    current_user=Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    from .version_models import ReportVersion
    # Get current report pages
    result = await db.execute(
        select(Report).where(Report.id == report_id, Report.tenant_id == current_user.tenant_id)
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado.")

    version = ReportVersion(
        report_id=report_id,
        tenant_id=current_user.tenant_id,
        pages=report.pages or [],
        title=report.title,
        label=label,
    )
    db.add(version)

    # Keep only last 20 versions
    old = await db.execute(
        select(ReportVersion)
        .where(ReportVersion.report_id == report_id, ReportVersion.tenant_id == current_user.tenant_id)
        .order_by(ReportVersion.created_at.desc())
        .offset(20)
    )
    for v in old.scalars().all():
        await db.delete(v)

    await db.commit()
    await db.refresh(version)
    return {"id": str(version.id), "created_at": version.created_at.isoformat()}


@router.post("/{report_id}/versions/{version_id}/restore")
async def restore_version(
    report_id: uuid.UUID,
    version_id: uuid.UUID,
    current_user=Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    from .version_models import ReportVersion
    from sqlalchemy.orm.attributes import flag_modified

    version = await db.scalar(
        select(ReportVersion).where(
            ReportVersion.id == version_id,
            ReportVersion.report_id == report_id,
            ReportVersion.tenant_id == current_user.tenant_id,
        )
    )
    if not version:
        raise HTTPException(status_code=404, detail="Versão não encontrada.")

    report = await db.scalar(
        select(Report).where(Report.id == report_id, Report.tenant_id == current_user.tenant_id)
    )
    if not report:
        raise HTTPException(status_code=404)

    report.pages = version.pages
    flag_modified(report, 'pages')
    await db.commit()
    return {"ok": True}


class ScheduledReportCreate(BaseModel):
    report_id: uuid.UUID
    report_name: str
    frequency: str  # daily | weekly | monthly
    emails: list[str]


class ScheduledReportUpdate(BaseModel):
    frequency: str | None = None
    emails: list[str] | None = None
    is_active: bool | None = None


class ScheduledReportResponse(BaseModel):
    id: uuid.UUID
    report_id: uuid.UUID
    report_name: str
    frequency: str
    emails: list[str]
    is_active: bool
    next_run: str
    last_run: str | None
    created_at: str

    model_config = {"from_attributes": True}


def _sr_to_response(sr: ScheduledReport) -> ScheduledReportResponse:
    return ScheduledReportResponse(
        id=sr.id,
        report_id=sr.report_id,
        report_name=sr.report_name,
        frequency=sr.frequency,
        emails=sr.emails or [],
        is_active=sr.is_active,
        next_run=sr.next_run.isoformat(),
        last_run=sr.last_run.isoformat() if sr.last_run else None,
        created_at=sr.created_at.isoformat(),
    )


@router.get(
    "/scheduled",
    response_model=list[ScheduledReportResponse],
    summary="Lista relatórios agendados do tenant",
)
async def list_scheduled_reports(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    svc = ScheduledReportService(db)
    items = await svc.list(current_user.tenant_id)
    return [_sr_to_response(sr) for sr in items]


@router.post(
    "/scheduled",
    response_model=ScheduledReportResponse,
    status_code=201,
    summary="Cria agendamento de relatório",
)
async def create_scheduled_report(
    data: ScheduledReportCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    valid_frequencies = {"daily", "weekly", "monthly"}
    if data.frequency not in valid_frequencies:
        raise HTTPException(
            status_code=422,
            detail=f"Frequência inválida. Use: {', '.join(valid_frequencies)}",
        )
    if not data.emails:
        raise HTTPException(status_code=422, detail="Informe pelo menos um email destinatário.")
    svc = ScheduledReportService(db)
    sr = await svc.create(
        tenant_id=current_user.tenant_id,
        report_id=data.report_id,
        report_name=data.report_name,
        frequency=data.frequency,
        emails=data.emails,
    )
    return _sr_to_response(sr)


@router.post(
    "/scheduled/run-due",
    summary="Dispara envio dos relatórios agendados com next_run vencido (para testes / cron externo)",
)
async def run_due_scheduled_reports(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Busca todos os agendamentos ativos com next_run <= agora, envia email
    para cada destinatário e atualiza last_run / next_run.
    Em produção, este endpoint deve ser chamado por um cron job externo (Railway Cron / Celery Beat).
    """
    from app.core.email import send_admin_email

    svc = ScheduledReportService(db)
    due = await svc.get_due()
    sent_count = 0
    errors: list[str] = []

    for sr in due:
        dashboard_url = f"https://app.jarbis.cc/dashboards/{sr.report_id}"
        subject = f"Relatório agendado: {sr.report_name}"
        html_body = (
            "<!DOCTYPE html>"
            '<html lang="pt-BR"><head><meta charset="utf-8">'
            f"<title>{subject}</title></head>"
            '<body style="font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',Helvetica,Arial,sans-serif;'
            'background:#f1f5f9;margin:0;padding:40px 20px;">'
            '<div style="max-width:480px;margin:0 auto;background:white;border-radius:16px;'
            'overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.08);">'
            '<div style="background:linear-gradient(135deg,#6D28D9 0%,#7C3AED 100%);padding:28px 32px;text-align:center;">'
            '<span style="color:white;font-weight:800;font-size:20px;">jarbis</span></div>'
            '<div style="padding:40px 32px 32px;">'
            '<h2 style="color:#0f172a;font-size:22px;font-weight:800;margin:0 0 12px;">'
            f"Seu relatório <b>{sr.report_name}</b> está pronto</h2>"
            '<p style="color:#475569;font-size:15px;line-height:1.6;margin:0 0 28px;">'
            f"Este é um envio automático agendado ({sr.frequency})."
            " Clique abaixo para visualizar o dashboard atualizado.</p>"
            '<div style="text-align:center;margin-bottom:28px;">'
            f'<a href="{dashboard_url}" style="display:inline-block;'
            "background:linear-gradient(135deg,#6D28D9 0%,#7C3AED 100%);"
            'color:white;font-size:15px;font-weight:700;padding:16px 32px;'
            'border-radius:12px;text-decoration:none;">'
            "Ver Dashboard &rarr;</a></div></div>"
            '<div style="border-top:1px solid #e2e8f0;padding:20px 32px;text-align:center;background:#f8fafc;">'
            '<p style="color:#94a3b8;font-size:12px;margin:0;">'
            '<a href="https://jarbis.cc" style="color:#94a3b8;text-decoration:none;">jarbis.cc</a>'
            " &nbsp;&middot;&nbsp; &copy; 2026 Jarbis</p></div></div></body></html>"
        )
        for email in sr.emails:
            try:
                await send_admin_email(email, subject, html_body)
                sent_count += 1
            except Exception as exc:
                errors.append(f"{email}: {exc}")
        await svc.mark_sent(sr)

    return {
        "processed": len(due),
        "emails_sent": sent_count,
        "errors": errors,
    }


@router.put(
    "/scheduled/{schedule_id}",
    response_model=ScheduledReportResponse,
    summary="Atualiza agendamento de relatório",
)
async def update_scheduled_report(
    schedule_id: uuid.UUID,
    data: ScheduledReportUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    if data.frequency is not None and data.frequency not in {"daily", "weekly", "monthly"}:
        raise HTTPException(status_code=422, detail="Frequência inválida. Use: daily, weekly, monthly")
    svc = ScheduledReportService(db)
    sr = await svc.update(
        schedule_id=schedule_id,
        tenant_id=current_user.tenant_id,
        frequency=data.frequency,
        emails=data.emails,
        is_active=data.is_active,
    )
    if not sr:
        raise HTTPException(status_code=404, detail="Agendamento não encontrado.")
    return _sr_to_response(sr)


@router.delete(
    "/scheduled/{schedule_id}",
    status_code=204,
    summary="Remove agendamento de relatório",
)
async def delete_scheduled_report(
    schedule_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    svc = ScheduledReportService(db)
    deleted = await svc.delete(schedule_id, current_user.tenant_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Agendamento não encontrado.")


# ---------------------------------------------------------------------------
# BI Advisor — diagnóstico inteligente de dashboard
# ---------------------------------------------------------------------------


@router.post(
    "/{report_id}/diagnose",
    summary="Diagnóstico inteligente do dashboard — insights, gaps e histórico De/Para",
)
async def diagnose_dashboard_endpoint(
    report_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    effective_tenant_id: uuid.UUID = Depends(get_effective_tenant_id),
):
    import json
    import os
    import re as _re

    import anthropic as ant

    from .ai_usage_models import AIUsageLog
    from .diagnosis_models import DiagnosisSnapshot

    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    plan = tenant.plan if tenant else "free"
    check_feature_allowed("ai", plan)
    await check_ai_query_limit(db, current_user.tenant_id, plan)

    # Carregar o report
    report = await db.scalar(
        select(Report).where(Report.id == report_id, Report.tenant_id == effective_tenant_id)
    )
    if not report:
        raise HTTPException(status_code=404, detail="Dashboard não encontrado")

    # Extrair blocos das páginas
    all_blocks = []
    pages = report.pages or []
    for page in pages:
        all_blocks.extend(page.get("blocks", []))
    if not all_blocks:
        all_blocks = report.blocks or []

    dataset_ids = list({b.get("dataset_id") for b in all_blocks if b.get("dataset_id")})

    if not dataset_ids:
        return {
            "domain": "generic", "domain_name": "Genérico",
            "visual_insights": ["O dashboard ainda não tem blocos conectados a dados."],
            "technical": {"total_rows": 0, "total_cols": 0, "block_types": {}, "missing_columns": []},
            "missing_blocks": [],
            "suggestions": ["Conecte um dataset e clique em Gerar com IA para criar os primeiros blocos."],
            "health_score": 10,
            "previous": None,
        }

    # Dataset principal
    svc = DatasetService(db)
    ds = None
    for did in dataset_ids:
        try:
            ds = await svc.get(uuid.UUID(did), effective_tenant_id)
            if not ds and effective_tenant_id != current_user.tenant_id:
                ds = await svc.get(uuid.UUID(did), current_user.tenant_id)
            if ds:
                break
        except Exception:
            continue

    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")

    columns = ds.columns or []

    # Rows: Redis (warp cache) > DB
    rows = ds.rows or []
    if not rows:
        try:
            from app.core.cache import get_redis as _get_redis
            _redis = _get_redis()
            _cached = await warp_get_rows(_redis, str(ds.id))
            if _cached:
                rows = _cached
        except Exception:
            pass

    total_rows = len(rows)
    domain_key, domain_tpl = _detect_domain(columns, rows)

    # Estatísticas das colunas
    col_stats: dict[str, dict] = {}
    for col in columns:
        vals = [r.get(col) for r in rows if r.get(col) is not None
                and str(r.get(col)).strip().lower() not in (
                    "", "null", "none", "n/a", "na", "n.a.", "-", "--", "nan",
                    "#n/a", "#na", "#div/0!", "#ref!", "#value!", "#num!", "#name?", "#null!",
                )]
        null_pct = round((total_rows - len(vals)) / total_rows * 100) if total_rows else 100
        nums = [float(v) for v in vals if isinstance(v, (int, float))]
        if nums:
            col_stats[col] = {"type": "numero", "sum": round(sum(nums), 2), "avg": round(sum(nums)/len(nums), 2), "null_pct": null_pct}
        else:
            col_stats[col] = {"type": "texto", "null_pct": null_pct}

    block_types: dict[str, int] = {}
    for b in all_blocks:
        bt = b.get("type", "unknown")
        block_types[bt] = block_types.get(bt, 0) + 1

    missing_cols_from_template = [
        imp for imp in domain_tpl.get("optional_improvements", [])
        if not any(imp["col"].lower().replace("_", "") in c.lower().replace("_", "") for c in columns)
    ]

    schema_lines = []
    for col, st in col_stats.items():
        if st["type"] == "numero":
            schema_lines.append(f"  {col}: número | total={st['sum']} | média={st['avg']}")
        else:
            schema_lines.append(f"  {col}: texto")

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key or api_key == "your_key_here":
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY não configurada.")

    system_prompt = (
        "Você é um analista de negócios especializado em BI para PMEs brasileiras. "
        "Responda APENAS com JSON válido, sem markdown.\n\n"
        "Estrutura OBRIGATÓRIA do JSON:\n"
        "{\n"
        '  "visual_insights": ["frase de negócio 1", "frase de negócio 2", "frase de negócio 3"],\n'
        '  "missing_blocks": [{"type": "line", "title": "Título", "reason": "Por que adicionar"}],\n'
        '  "missing_columns": [{"col": "NOME", "impact": "O que permitiria analisar"}],\n'
        '  "suggestions": ["ação concreta 1", "ação concreta 2"],\n'
        '  "health_score": 75\n'
        "}\n\n"
        "REGRAS:\n"
        "- visual_insights: 2-3 frases sobre o NEGÓCIO (faturamento, clientes, tendências). "
        "  Use os números reais do schema. PROIBIDO mencionar 'banco de dados', 'colunas', 'blocos', 'dataset'. "
        "  Fale como analista para o dono da empresa, não para um técnico.\n"
        "- missing_blocks: até 2 gráficos que faltam. Se já está completo, lista vazia.\n"
        "- missing_columns: até 2 colunas que agregariam valor de negócio. Se já está bom, lista vazia.\n"
        "- suggestions: 1-2 ações de negócio que o usuário pode tomar com os dados.\n"
        "- health_score: 0-100. 100 = dashboard completo para o domínio detectado.\n"
        "- Responda em português brasileiro, linguagem simples e direta."
    )

    user_msg = (
        f"SETOR/DOMÍNIO: {domain_tpl['name']}\n"
        f"DADOS: {total_rows} registros\n\n"
        f"MÉTRICAS DISPONÍVEIS:\n" + "\n".join(schema_lines or [f"  {c}" for c in columns[:12]]) + "\n\n"
        f"VISUALIZAÇÕES NO DASHBOARD: {json.dumps(block_types, ensure_ascii=False)}\n"
        + (f"\nCOLUNAS EXTRAS QUE PODERIAM EXISTIR:\n"
           + "\n".join(f"  {m['col']}: {m['impact']}" for m in missing_cols_from_template[:2])
           if missing_cols_from_template else "")
        + "\n\nGere o diagnóstico:"
    )

    client = ant.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1024,
        system=system_prompt,
        messages=[{"role": "user", "content": user_msg}],
    )

    # Log de uso de IA
    try:
        usage = msg.usage
        log = AIUsageLog(
            tenant_id=current_user.tenant_id, user_id=current_user.id, dataset_id=ds.id,
            model="claude-haiku-4-5-20251001",
            tokens_input=usage.input_tokens, tokens_output=usage.output_tokens,
            question="[diagnose-dashboard]",
        )
        db.add(log)
    except Exception:
        pass

    text = msg.content[0].text.strip()
    try:
        m2 = _re.search(r"\{.*\}", text, _re.DOTALL)
        ai_result = json.loads(m2.group() if m2 else text)
    except Exception:
        ai_result = {
            "visual_insights": [f"Dashboard com {len(all_blocks)} visualizações no domínio {domain_tpl['name']}."],
            "missing_blocks": [], "missing_columns": [], "suggestions": [], "health_score": 50,
        }

    visual_insights = ai_result.get("visual_insights") or ai_result.get("insights") or []
    missing_blocks_ai = ai_result.get("missing_blocks", [])[:2]
    missing_columns_ai = ai_result.get("missing_columns", [])[:2]
    suggestions_ai = ai_result.get("suggestions", [])[:2]
    health_score = int(ai_result.get("health_score", 50))

    # Dados técnicos (separados dos visuais)
    technical = {
        "total_rows": total_rows,
        "total_cols": len(columns),
        "block_types": block_types,
        "missing_columns": missing_columns_ai,
    }

    # Carregar snapshot anterior para comparação De/Para
    from datetime import timezone as _tz
    prev_snapshot = await db.scalar(
        select(DiagnosisSnapshot)
        .where(DiagnosisSnapshot.report_id == report_id, DiagnosisSnapshot.tenant_id == current_user.tenant_id)
        .order_by(DiagnosisSnapshot.created_at.desc())
        .limit(1)
    )
    previous = None
    if prev_snapshot:
        previous = {
            "health_score": prev_snapshot.health_score,
            "visual_insights": prev_snapshot.visual_insights,
            "created_at": prev_snapshot.created_at.isoformat(),
            "delta": health_score - prev_snapshot.health_score,
        }

    # Salvar novo snapshot
    snapshot = DiagnosisSnapshot(
        report_id=report_id,
        tenant_id=current_user.tenant_id,
        health_score=health_score,
        domain=domain_key,
        domain_name=domain_tpl["name"],
        visual_insights=visual_insights,
        missing_blocks=missing_blocks_ai,
        missing_columns=missing_columns_ai,
        suggestions=suggestions_ai,
    )
    db.add(snapshot)
    await db.commit()

    return {
        "domain": domain_key,
        "domain_name": domain_tpl["name"],
        "visual_insights": visual_insights,
        "missing_blocks": missing_blocks_ai,
        "technical": technical,
        "suggestions": suggestions_ai,
        "health_score": health_score,
        "previous": previous,
    }


@router.post(
    "/{report_id}/suggest-blocks",
    summary="Sugere novos blocos para colunas ainda não visualizadas no dashboard",
)
async def suggest_blocks_endpoint(
    report_id: uuid.UUID,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    effective_tenant_id: uuid.UUID = Depends(get_effective_tenant_id),
):
    import json
    import os
    import re as _re
    import uuid as _uuid

    import anthropic as ant

    from .ai_usage_models import AIUsageLog

    tenant = await db.scalar(select(Tenant).where(Tenant.id == current_user.tenant_id))
    plan = tenant.plan if tenant else "free"
    check_feature_allowed("ai", plan)
    await check_ai_query_limit(db, current_user.tenant_id, plan)

    # Carregar report
    report = await db.scalar(
        select(Report).where(Report.id == report_id, Report.tenant_id == effective_tenant_id)
    )
    if not report:
        raise HTTPException(status_code=404, detail="Dashboard não encontrado")

    # Dataset ID do body
    dataset_id_raw = data.get("dataset_id")
    if not dataset_id_raw:
        raise HTTPException(status_code=422, detail="dataset_id obrigatório")
    try:
        dataset_id = uuid.UUID(str(dataset_id_raw))
    except ValueError:
        raise HTTPException(status_code=422, detail="dataset_id inválido")

    svc = DatasetService(db)
    ds = await svc.get(dataset_id, effective_tenant_id)
    if not ds and effective_tenant_id != current_user.tenant_id:
        ds = await svc.get(dataset_id, current_user.tenant_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Dataset não encontrado")

    columns = ds.columns or []
    rows = ds.rows or []
    if not rows:
        try:
            from app.core.cache import get_redis as _get_redis
            _redis = _get_redis()
            _cached = await warp_get_rows(_redis, str(ds.id))
            if _cached:
                rows = _cached
        except Exception:
            pass

    # Extrair colunas já usadas nos blocos existentes
    all_blocks = []
    pages = report.pages or []
    for page in pages:
        all_blocks.extend(page.get("blocks", []))
    if not all_blocks:
        all_blocks = report.blocks or []

    used_cols: set[str] = set()
    for b in all_blocks:
        if b.get("label_col"):
            used_cols.add(b["label_col"])
        if b.get("value_col"):
            used_cols.add(b["value_col"])
        if b.get("filter_col"):
            used_cols.add(b["filter_col"])

    unused_cols = [c for c in columns if c not in used_cols]

    if not unused_cols:
        return {"blocks": [], "message": "Todas as colunas já estão visualizadas.", "new_cols": []}

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key or api_key == "your_key_here":
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY não configurada.")

    # Estatísticas das colunas não utilizadas
    total_rows = len(rows)
    col_stats: dict[str, dict] = {}
    for col in unused_cols:
        vals = [r.get(col) for r in rows if r.get(col) is not None
                and str(r.get(col)).strip().lower() not in ("", "null", "none", "n/a", "na", "-", "--", "nan")]
        if not vals:
            col_stats[col] = {"type": "texto", "null_pct": 100}
            continue
        nums: list[float] = []
        all_numeric = True
        for v in vals:
            try:
                nums.append(float(str(v).replace(",", ".")))
            except (ValueError, TypeError):
                all_numeric = False
                break
        if all_numeric and nums:
            nonzero = sum(1 for n in nums if n != 0)
            null_pct = round((total_rows - len(vals)) / total_rows * 100) if total_rows else 0
            nonzero_pct = round(nonzero / len(nums) * 100)
            flag = "★ INTERESSANTE" if nonzero_pct >= 20 and null_pct <= 70 else "— zero/irrelevante"
            col_stats[col] = {"type": "numero", "sum": round(sum(nums), 2), "flag": flag,
                               "nonzero_pct": nonzero_pct, "null_pct": null_pct}
        else:
            import re as _re2
            sample_val = str(vals[0])
            if _re2.match(r"\d{4}-\d{2}|\d{2}/\d{4}|\d{2}/\d{2}/\d{4}", sample_val):
                col_stats[col] = {"type": "data", "sample": sample_val}
            else:
                unique_vals = list(dict.fromkeys([str(v) for v in vals]))
                col_stats[col] = {"type": "texto", "unique": len(set(str(v) for v in vals)), "top3": unique_vals[:3]}

    # Schema das colunas novas
    schema_parts = [f"Dataset: {ds.name}", f"Novas colunas a visualizar ({len(unused_cols)} de {len(columns)} total):"]
    has_interesting = False
    for col, st in col_stats.items():
        if st["type"] == "numero":
            flag = st.get("flag", "")
            schema_parts.append(f'  "{col}" [número {flag}] sum={st["sum"]}, nonzero={st.get("nonzero_pct")}%')
            if "INTERESSANTE" in flag:
                has_interesting = True
        elif st["type"] == "data":
            schema_parts.append(f'  "{col}" [data] ex: {st["sample"]}')
            has_interesting = True
        else:
            top3 = ", ".join(st.get("top3", []))
            schema_parts.append(f'  "{col}" [texto] {st.get("unique","?")} valores únicos, ex: {top3}')
            if st.get("unique", 0) > 1:
                has_interesting = True

    if not has_interesting:
        return {"blocks": [], "message": "As colunas disponíveis não têm dados suficientes para novas visualizações.", "new_cols": unused_cols}

    schema_str = "\n".join(schema_parts)

    system_prompt = (
        "Você é um especialista em Business Intelligence para PMEs brasileiras.\n"
        "Analise as NOVAS colunas informadas e gere blocos de dashboard para elas.\n\n"
        "RETORNE SOMENTE um JSON array válido (sem markdown, sem texto fora do JSON):\n"
        '[\n'
        '  {"type":"kpi","title":"Título","value_col":"coluna","agg":"sum","label_col":null,"layout":{"w":3,"h":2}},\n'
        '  {"type":"bar","title":"Título","label_col":"coluna_texto","value_col":"coluna_numero","agg":"sum","layout":{"w":6,"h":4}}\n'
        ']\n\n'
        "TIPOS: kpi, bar, bar_h, line, area, pie\n"
        "TAMANHOS: kpi → w=3,h=2 | line/area/bar/bar_h → w=6,h=4 | pie → w=4,h=4\n\n"
        "REGRAS:\n"
        "1. Use APENAS as colunas listadas (★ INTERESSANTE)\n"
        "2. Ignore colunas '— zero/irrelevante' e null_pct>70%\n"
        "3. Títulos em português, máx 35 chars\n"
        "4. Cada bloco responde UMA pergunta de negócio\n"
        "5. Se não há colunas interessantes, retorne []\n"
        "6. Alíquotas e percentuais: agg=avg"
    )

    client = ant.Anthropic(api_key=api_key)
    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1024,
        system=system_prompt,
        messages=[{"role": "user", "content": f"{schema_str}\n\nGere os blocos JSON:"}],
    )

    # Log de uso
    try:
        usage = msg.usage
        log = AIUsageLog(
            tenant_id=current_user.tenant_id, user_id=current_user.id, dataset_id=dataset_id,
            model="claude-haiku-4-5-20251001",
            tokens_input=usage.input_tokens, tokens_output=usage.output_tokens,
            question="[suggest-blocks]",
        )
        db.add(log)
        await db.commit()
    except Exception:
        pass

    text = msg.content[0].text.strip()
    try:
        blocks = json.loads(text)
    except json.JSONDecodeError:
        m = _re.search(r"\[.*\]", text, _re.DOTALL)
        blocks = json.loads(m.group()) if m else []

    if not isinstance(blocks, list):
        blocks = []

    _layout_defaults = {"kpi": {"w": 3, "h": 2}, "line": {"w": 6, "h": 4}, "area": {"w": 6, "h": 4},
                        "bar": {"w": 6, "h": 4}, "bar_h": {"w": 6, "h": 4}, "pie": {"w": 4, "h": 4}}
    for b in blocks:
        b["dataset_id"] = str(dataset_id)
        if not b.get("id"):
            b["id"] = str(_uuid.uuid4())
        btype = b.get("type", "bar")
        default_size = _layout_defaults.get(btype, {"w": 6, "h": 4})
        if not isinstance(b.get("layout"), dict):
            b["layout"] = default_size.copy()
        else:
            b["layout"].setdefault("w", default_size["w"])
            b["layout"].setdefault("h", default_size["h"])

    msg_text = f"{len(blocks)} novo(s) bloco(s) gerado(s) para {len(unused_cols)} coluna(s) não visualizada(s)." if blocks else "Nenhum bloco gerado para as colunas disponíveis."
    return {"blocks": blocks, "message": msg_text, "new_cols": unused_cols}


# ---------------------------------------------------------------------------
# Report CRUD — parametrizado /{report_id} deve ficar por ÚLTIMO para não
# capturar rotas fixas como /datasets, /public, /alerts, /ai-query, /data
# ---------------------------------------------------------------------------


@router.get(
    "/{report_id}",
    response_model=ReportResponse,
    summary="Detalhe do relatório",
)
async def get_report(
    report_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = ReportService(db)
    report = await service.get(report_id, current_user.tenant_id)
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")
    return report


@router.put(
    "/{report_id}",
    response_model=ReportResponse,
    summary="Atualiza título, descrição ou blocos do relatório",
)
async def update_report(
    report_id: uuid.UUID,
    data: ReportUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = ReportService(db)
    report = await service.update(report_id, current_user.tenant_id, data)
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")
    return report


@router.delete(
    "/{report_id}",
    status_code=204,
    summary="Remove o relatório",
)
async def delete_report(
    report_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    service = ReportService(db)
    deleted = await service.delete(report_id, current_user.tenant_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")


@router.post(
    "/{report_id}/share",
    response_model=ShareResponse,
    summary="Gera link público de compartilhamento",
)
async def create_share_link(
    report_id: uuid.UUID,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Cria (ou retorna) um link público para o relatório ser visualizado sem login.
    A URL do frontend é gerada substituindo a porta 8000 pela 3000 e usando
    o caminho /r/{token}.
    """
    base_url = str(request.base_url).rstrip("/")
    service = ReportService(db)
    result = await service.get_or_create_share(report_id, current_user.tenant_id, base_url)
    if not result:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")
    return ShareResponse(**result)
