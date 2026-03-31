"""Service — Datasets de Relatórios."""

from __future__ import annotations

import csv
import io
import ipaddress
import json
import socket
import uuid
from datetime import date, datetime, timezone
from typing import Any
from urllib.parse import urlparse

import httpx
import openpyxl
from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from .dataset_models import ReportDataset


_BLOCKED_HOSTS = {
    "localhost",
    "metadata.google.internal",
    "169.254.169.254",  # AWS/GCP metadata
    "instance-data",
}


def _validate_ssrf(url: str) -> None:
    """Bloqueia URLs internas/privadas para prevenir SSRF."""
    try:
        parsed = urlparse(url)
    except Exception:
        raise HTTPException(status_code=400, detail="URL inválida.")

    if parsed.scheme not in ("http", "https"):
        raise HTTPException(status_code=400, detail="Apenas URLs http/https são permitidas.")

    hostname = parsed.hostname
    if not hostname:
        raise HTTPException(status_code=400, detail="URL sem hostname.")

    if hostname.lower() in _BLOCKED_HOSTS:
        raise HTTPException(status_code=400, detail="URL interna não permitida.")

    # Tenta parsear como IP direto; caso seja hostname, resolve
    try:
        addr = ipaddress.ip_address(hostname)
    except ValueError:
        try:
            addr = ipaddress.ip_address(socket.gethostbyname(hostname))
        except Exception:
            raise HTTPException(status_code=400, detail=f"Não foi possível resolver o hostname: {hostname}")

    if (
        addr.is_private
        or addr.is_loopback
        or addr.is_link_local
        or addr.is_reserved
        or addr.is_multicast
    ):
        raise HTTPException(status_code=400, detail="URLs de redes privadas não são permitidas.")


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _detect_columns(rows: list[dict]) -> list[str]:
    if not rows:
        return []
    return list(rows[0].keys())


_NULL_STRINGS = frozenset({
    "", "null", "none", "n/a", "na", "n.a.", "-", "--", "nan",
    "#n/a", "#na", "#div/0!", "#ref!", "#value!", "#num!", "#name?", "#null!",
})


_CURRENCY_PREFIX_RE = None  # inicializado lazy


def _try_parse_number(s: str) -> float | None:
    """Tenta parsear string como número, suportando formatos internacionais e monetários.

    Suporta:
    - Formato BR: "R$ 11.174,44" → 11174.44
    - Formato US: "$1,234.56" → 1234.56
    - Só vírgula decimal: "11,5" → 11.5
    - Percentual: "5,25%" → 5.25
    - Sem prefixo: "11174.44" → 11174.44
    """
    import re as _re
    # Remove símbolos de moeda e espaços: R$, $, €, £, ¥, etc.
    cleaned = _re.sub(r'[R\$€£¥\s]', '', s).strip()
    # Remove sinal de percentual (guarda para aplicar depois se necessário)
    cleaned = cleaned.rstrip('%')
    if not cleaned:
        return None
    # Formato: tem vírgula E ponto → vírgula é decimal, ponto é milhar (BR/EU)
    #          "11.174,44" → "11174.44"
    if ',' in cleaned and '.' in cleaned:
        # Verifica qual vem por último para decidir qual é o decimal
        last_comma = cleaned.rfind(',')
        last_dot = cleaned.rfind('.')
        if last_comma > last_dot:
            # Ponto é milhar, vírgula é decimal: 1.234,56 → 1234.56
            cleaned = cleaned.replace('.', '').replace(',', '.')
        else:
            # Vírgula é milhar, ponto é decimal: 1,234.56 → 1234.56
            cleaned = cleaned.replace(',', '')
    elif ',' in cleaned:
        # Só vírgula — é separador decimal: "11,5" → "11.5"
        cleaned = cleaned.replace(',', '.')
    # Caso só tenha ponto: deixa como está (decimal padrão internacional)
    try:
        return float(cleaned)
    except ValueError:
        return None


def _coerce_row(row: dict) -> dict:
    """Tenta converter strings numéricas para float. Normaliza representações comuns de nulo.

    Trata especificamente:
    - Formato monetário BR: "R$ 11.174,44" → 11174.44
    - Erros de fórmula Excel (CellErrorValue: #N/A, #REF!, #DIV/0!, …) → None
    - float NaN/Inf → None
    - Qualquer tipo não reconhecido → str → testa null
    """
    import math

    out = {}
    for k, v in row.items():
        if v is None:
            out[k] = None
        elif isinstance(v, bool):
            out[k] = v
        elif isinstance(v, (date, datetime)):
            out[k] = v  # datas preservadas (JSON-serializable via isoformat)
        elif isinstance(v, (int, float)):
            f = float(v)
            out[k] = None if (math.isnan(f) or math.isinf(f)) else v
        elif isinstance(v, str):
            stripped = v.strip()
            if stripped.lower() in _NULL_STRINGS:
                out[k] = None
                continue
            parsed = _try_parse_number(stripped)
            if parsed is not None:
                out[k] = parsed
            else:
                out[k] = stripped if stripped else None
        else:
            # CellErrorValue do openpyxl (#N/A, #REF!, #DIV/0!, etc.) e outros tipos
            s = str(v).strip()
            out[k] = None if (s.lower() in _NULL_STRINGS or s.startswith("#")) else s
    return out


def _apply_computed_columns(rows: list[dict], computed_columns: list[dict]) -> list[dict]:
    """Aplica colunas calculadas a cada row. computed_columns: [{name, expression, refs}]"""
    if not computed_columns or not rows:
        return rows
    try:
        from simpleeval import simple_eval  # noqa: F401 — verifica disponibilidade
    except ImportError:
        return rows

    result = []
    for row in rows:
        new_row = dict(row)
        for col in computed_columns:
            name = col.get("name", "").strip()
            expr = col.get("expression", "").strip()
            if not name or not expr:
                continue
            try:
                names_num: dict[str, Any] = {}
                for k, v in new_row.items():
                    if isinstance(v, (int, float)):
                        names_num[k] = v
                    elif isinstance(v, str):
                        try:
                            names_num[k] = float(v.replace(",", "."))
                        except ValueError:
                            names_num[k] = 0
                    elif v is None:
                        names_num[k] = 0
                from simpleeval import simple_eval
                val = simple_eval(expr, names=names_num)
                new_row[name] = round(float(val), 4) if isinstance(val, (int, float)) else val
            except Exception:
                new_row[name] = None
        result.append(new_row)
    return result


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [_coerce_row(dict(row)) for row in reader]


def _count_non_null(rows: list[tuple]) -> int:
    """Conta valores não-nulos em uma lista de linhas."""
    return sum(1 for r in rows for v in r if v is not None)


def _resolve_formula_sheet(content: bytes, sheet_name: str) -> str | None:
    """Se a aba é um espelho por fórmula (sem dados cached), detecta e retorna a aba fonte.

    Abre o arquivo sem data_only para ler as fórmulas brutas. Detecta padrões como:
    - =ARRAYFORMULA(BASE_DADOS!A1:BU)
    - =BASE_DADOS!A1
    - ='Aba Dados'!B2
    """
    import re as _re
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=False)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[sheet_name]
        # Verifica as primeiras 5 linhas / 10 células em busca de referências cross-sheet
        for row in ws.iter_rows(max_row=5, values_only=False):
            for cell in row:
                val = cell.value
                if val and isinstance(val, str) and val.startswith("="):
                    # Padrões: =FUNC('Sheet'!...) ou =FUNC(Sheet!...) ou ='Sheet'!A1
                    m = _re.search(r"[=(,]'?([^'!\(\),\s][^'!\(\),]*?)'?!", val)
                    if m:
                        ref = m.group(1).strip()
                        if ref in wb.sheetnames and ref != sheet_name:
                            wb.close()
                            return ref
        wb.close()
    except Exception:
        pass
    return None


def _find_header_row(rows: list[tuple]) -> int:
    """Encontra o índice da linha de cabeçalho real.

    Ignora linhas de título/metadados no topo da aba — linhas onde quase tudo
    é None ou há apenas 1-2 células preenchidas. Retorna o índice da primeira
    linha com pelo menos 3 valores não-nulos e maioria de strings (cabeçalhos).
    """
    for i, row in enumerate(rows[:20]):
        non_null = [v for v in row if v is not None]
        if len(non_null) < 2:
            continue
        # Linha de cabeçalho: maioria são strings (não números puros)
        string_count = sum(1 for v in non_null if isinstance(v, str))
        if string_count >= max(2, len(non_null) * 0.5):
            return i
    return 0  # fallback: primeira linha


def _parse_excel(content: bytes, sheet_name: str | None = None) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    actual_name = ws.title
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Se a aba parece vazia (fórmulas sem cache), tenta detectar aba fonte
    if _count_non_null(rows[:10]) < 3 and len(rows) > 1:
        source = _resolve_formula_sheet(content, actual_name)
        if source:
            wb2 = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            if source in wb2.sheetnames:
                rows = list(wb2[source].iter_rows(values_only=True))
            wb2.close()

    if not rows:
        return []
    header_idx = _find_header_row(rows)
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[header_idx])]
    result = []
    for row in rows[header_idx + 1:]:
        if all(v is None for v in row):
            continue
        result.append(_coerce_row({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}))
    return result


def _extract_json_path(data: Any, path: str | None) -> list:
    """Navega pelo JSON usando notação de ponto. Ex: 'data.items'"""
    if not path:
        if isinstance(data, list):
            return data
        return [data] if isinstance(data, dict) else []
    parts = path.split(".")
    current = data
    for part in parts:
        if isinstance(current, dict):
            current = current.get(part)
        else:
            break
    if isinstance(current, list):
        return current
    return [current] if isinstance(current, dict) else []


def _parse_date_value(v: Any) -> date | None:
    """Tenta parsear um valor como date. Suporta ISO, BR (DD/MM/YYYY) e variantes."""
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _aggregate(rows: list[dict], label_col: str, value_col: str, agg: str) -> list[dict]:
    """Agrupa rows por label_col e agrega value_col.

    Células nulas (None) são IGNORADAS nas agregações sum/avg/max/min — não
    são tratadas como zero. Isso evita distorção em datasets com muitas células
    em branco: avg([None, None, None, 100]) = 100, não 25.
    """
    # groups[label] = (sum_valid, non_null_values, total_count)
    groups: dict[str, dict] = {}
    for row in rows:
        label = str(row.get(label_col, "")) if row.get(label_col) is not None else "(vazio)"
        raw = row.get(value_col)
        entry = groups.setdefault(label, {"vals": [], "total": 0})
        entry["total"] += 1
        if raw is None or str(raw).strip().lower() in ("", "null", "none", "n/a", "na", "-"):
            continue  # ignora nulo/vazio — não trata como zero
        try:
            entry["vals"].append(float(raw))
        except (TypeError, ValueError):
            pass  # valor não numérico também ignorado

    result = []
    for label, entry in groups.items():
        values = entry["vals"]
        total = entry["total"]
        if agg == "sum":
            v = sum(values)
        elif agg == "avg":
            v = sum(values) / len(values) if values else 0
        elif agg == "count":
            v = float(total)  # conta todas as linhas, incluindo nulas
        elif agg == "count_nonzero":
            v = float(sum(1 for x in values if x != 0))
        elif agg == "max":
            v = max(values) if values else 0
        elif agg == "min":
            v = min(values) if values else 0
        else:  # none — first non-null
            v = values[0] if values else 0
        result.append({"label": label, "value": round(v, 4)})

    return result


class DatasetService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def list(self, tenant_id: uuid.UUID) -> list[ReportDataset]:
        result = await self.db.execute(
            select(ReportDataset)
            .where(ReportDataset.tenant_id == tenant_id)
            .order_by(ReportDataset.created_at.desc())
        )
        return result.scalars().all()

    async def get(self, dataset_id: uuid.UUID, tenant_id: uuid.UUID) -> ReportDataset | None:
        result = await self.db.execute(
            select(ReportDataset).where(
                ReportDataset.id == dataset_id,
                ReportDataset.tenant_id == tenant_id,
            )
        )
        return result.scalar_one_or_none()

    async def create_from_file(
        self,
        tenant_id: uuid.UUID,
        name: str,
        filename: str,
        content: bytes,
        sheet_name: str | None = None,
        max_rows: int = -1,
    ) -> ReportDataset:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext in ("xlsx", "xls"):
            rows = _parse_excel(content, sheet_name=sheet_name)
            ds_type = "excel"
        else:
            rows = _parse_csv(content)
            ds_type = "csv"

        if max_rows != -1 and len(rows) > max_rows:
            raise HTTPException(
                status_code=402,
                detail=f"Seu plano suporta até {max_rows:,} linhas por dataset. Este arquivo tem {len(rows):,} linhas. Adicione um pack de linhas para continuar.".replace(",", ".")
            )

        columns = _detect_columns(rows)
        ds = ReportDataset(
            tenant_id=tenant_id,
            name=name,
            type=ds_type,
            rows=rows,
            columns=columns,
            row_count=len(rows),
        )
        self.db.add(ds)
        await self.db.commit()
        await self.db.refresh(ds)
        return ds

    async def create_from_api(
        self,
        tenant_id: uuid.UUID,
        name: str,
        api_url: str,
        api_headers: dict | None,
        api_data_path: str | None,
        max_rows: int = -1,
    ) -> ReportDataset:
        rows, columns = await self._fetch_api(api_url, api_headers or {}, api_data_path)
        if max_rows != -1 and len(rows) > max_rows:
            raise HTTPException(
                status_code=402,
                detail=f"Seu plano suporta até {max_rows:,} linhas por dataset. A API retornou {len(rows):,} linhas. Adicione um pack de linhas para continuar.".replace(",", ".")
            )
        ds = ReportDataset(
            tenant_id=tenant_id,
            name=name,
            type="api",
            rows=rows,
            columns=columns,
            row_count=len(rows),
            api_url=api_url,
            api_headers=api_headers,
            api_data_path=api_data_path,
            last_synced_at=_utcnow(),
        )
        self.db.add(ds)
        await self.db.commit()
        await self.db.refresh(ds)
        return ds

    async def sync_api(self, dataset_id: uuid.UUID, tenant_id: uuid.UUID, max_rows: int = -1) -> ReportDataset | None:
        ds = await self.get(dataset_id, tenant_id)
        if not ds or ds.type != "api" or not ds.api_url:
            return None
        new_rows, columns = await self._fetch_api(ds.api_url, ds.api_headers or {}, ds.api_data_path)
        if max_rows != -1 and len(new_rows) > max_rows:
            raise HTTPException(
                status_code=402,
                detail=f"Seu plano suporta até {max_rows:,} linhas por dataset. A API retornou {len(new_rows):,} linhas. Adicione um pack de linhas para continuar.".replace(",", ".")
            )

        sync_mode = getattr(ds, 'sync_mode', 'replace') or 'replace'
        if sync_mode == 'append':
            # Acumula linhas novas, evitando duplicatas por hash da linha
            existing = ds.rows or []
            existing_hashes = {hash(frozenset(str(r.items()))) for r in existing}
            added = [r for r in new_rows if hash(frozenset(str(r.items()))) not in existing_hashes]
            merged = existing + added
            from sqlalchemy.orm.attributes import flag_modified
            ds.rows = merged
            ds.columns = _detect_columns(merged)
            ds.row_count = len(merged)
            flag_modified(ds, 'rows')
            flag_modified(ds, 'columns')
        else:
            from sqlalchemy.orm.attributes import flag_modified
            ds.rows = new_rows
            ds.columns = columns
            ds.row_count = len(new_rows)
            flag_modified(ds, 'rows')
            flag_modified(ds, 'columns')

        ds.last_synced_at = _utcnow()
        await self.db.commit()
        await self.db.refresh(ds)
        return ds

    async def _fetch_links_data(
        self,
        tenant_id: uuid.UUID,
        campaign_id: str,
        days: int = 90,
    ) -> tuple[list[dict], list[str]]:
        """
        Busca dados de cliques do ClickHouse para uma campanha de links.
        Retorna (rows, columns) para materializar como dataset.
        Colunas: data, hora, dia_semana, campanha, link, ativo, dispositivo,
                 navegador, sistema, pais, cidade, fonte, cliques, cliques_unicos_periodo
        """
        from app.modules.links.models import ShortLink, LinkCampaign
        from app.config import settings as _settings
        from app.core.clickhouse import get_clickhouse
        import uuid as _uuid

        # Busca links da campanha (PostgreSQL)
        result = await self.db.scalars(
            select(ShortLink).where(
                ShortLink.campaign_id == _uuid.UUID(campaign_id),
                ShortLink.tenant_id == tenant_id,
            )
        )
        links = result.all()
        if not links:
            return [], []

        # Busca nome da campanha (PostgreSQL)
        campaign_obj = await self.db.scalar(
            select(LinkCampaign).where(LinkCampaign.id == _uuid.UUID(campaign_id))
        )
        campaign_name = campaign_obj.name if campaign_obj else "Desconhecida"

        link_ids = [str(lnk.id) for lnk in links]
        name_map = {str(lnk.id): lnk.name for lnk in links}
        active_map = {str(lnk.id): lnk.is_active for lnk in links}

        if not _settings.clickhouse_host:
            return [], []

        ids_str = ", ".join(f"'{lid}'" for lid in link_ids)
        ch = get_clickhouse()

        _DIAS = {1: 'Segunda', 2: 'Terça', 3: 'Quarta', 4: 'Quinta', 5: 'Sexta', 6: 'Sábado', 7: 'Domingo'}

        # Query 1: granular rows com hora e dia da semana
        rows_raw = ch.query(f"""
            SELECT
                formatDateTime(clicked_at, '%Y-%m-%d') as data,
                toHour(clicked_at) as hora,
                toDayOfWeek(clicked_at) as dia_semana_num,
                link_id,
                device_type,
                browser,
                os,
                country,
                city,
                multiIf(
                    referrer = '', 'Direto',
                    referrer LIKE '%google%', 'Google',
                    referrer LIKE '%instagram%', 'Instagram',
                    referrer LIKE '%facebook%' OR referrer LIKE '%fb.com%', 'Facebook',
                    referrer LIKE '%whatsapp%' OR referrer LIKE '%wa.me%', 'WhatsApp',
                    referrer LIKE '%t.co%' OR referrer LIKE '%twitter%' OR referrer LIKE '%x.com%', 'Twitter / X',
                    referrer LIKE '%linkedin%', 'LinkedIn',
                    referrer LIKE '%tiktok%', 'TikTok',
                    referrer LIKE '%youtube%' OR referrer LIKE '%youtu.be%', 'YouTube',
                    'Outro'
                ) as fonte,
                count() as cliques
            FROM link_events
            WHERE tenant_id = '{str(tenant_id)}'
              AND link_id IN ({ids_str})
              AND clicked_at >= now() - INTERVAL {days} DAY
            GROUP BY data, hora, dia_semana_num, link_id, device_type, browser, os, country, city, fonte
            ORDER BY data DESC, hora
        """).result_rows

        # Query 2: cliques únicos por link no período
        unicos_raw = ch.query(f"""
            SELECT
                link_id,
                countDistinct(ip_address) as cliques_unicos_periodo
            FROM link_events
            WHERE tenant_id = '{str(tenant_id)}'
              AND link_id IN ({ids_str})
              AND clicked_at >= now() - INTERVAL {days} DAY
            GROUP BY link_id
        """).result_rows
        unicos_map = {str(r[0]): int(r[1]) for r in unicos_raw}

        columns = [
            "data", "hora", "dia_semana", "campanha", "link", "ativo",
            "dispositivo", "navegador", "sistema", "pais", "cidade", "fonte",
            "cliques", "cliques_unicos_periodo",
        ]
        rows = [
            {
                "data": r[0],
                "hora": int(r[1]),
                "dia_semana": _DIAS.get(int(r[2]), ''),
                "campanha": campaign_name,
                "link": name_map.get(str(r[3]), str(r[3])),
                "ativo": "Sim" if active_map.get(str(r[3]), True) else "Não",
                "dispositivo": r[4] or "Outro",
                "navegador": r[5] or "Outro",
                "sistema": r[6] or "Outro",
                "pais": r[7] or "Desconhecido",
                "cidade": r[8] or "Desconhecido",
                "fonte": r[9],
                "cliques": int(r[10]),
                "cliques_unicos_periodo": unicos_map.get(str(r[3]), 0),
            }
            for r in rows_raw
        ]
        return rows, columns

    async def create_from_links_campaign(
        self,
        tenant_id: uuid.UUID,
        campaign_id: str,
        name: str,
        days: int = 90,
    ) -> "ReportDataset":
        rows, columns = await self._fetch_links_data(tenant_id, campaign_id, days)
        ds = ReportDataset(
            tenant_id=tenant_id,
            name=name,
            type="links",
            rows=rows,
            columns=columns,
            row_count=len(rows),
            links_campaign_id=campaign_id,
            links_days=days,
            last_synced_at=_utcnow(),
        )
        self.db.add(ds)
        await self.db.commit()
        await self.db.refresh(ds)
        return ds

    async def sync_links_campaign(
        self,
        dataset_id: uuid.UUID,
        tenant_id: uuid.UUID,
    ) -> "ReportDataset | None":
        from sqlalchemy.orm.attributes import flag_modified
        ds = await self.get(dataset_id, tenant_id)
        if not ds or ds.type != "links" or not ds.links_campaign_id:
            return None
        days = ds.links_days or 90
        rows, columns = await self._fetch_links_data(tenant_id, ds.links_campaign_id, days)
        ds.rows = rows
        ds.columns = columns
        ds.row_count = len(rows)
        ds.last_synced_at = _utcnow()
        flag_modified(ds, "rows")
        flag_modified(ds, "columns")
        await self.db.commit()
        await self.db.refresh(ds)
        return ds

    async def delete(self, dataset_id: uuid.UUID, tenant_id: uuid.UUID) -> bool:
        ds = await self.get(dataset_id, tenant_id)
        if not ds:
            return False
        await self.db.delete(ds)
        await self.db.commit()
        return True

    _VALID_AGGS = {"sum", "avg", "count", "max", "min", "none"}

    def query(
        self,
        ds: ReportDataset,
        label_col: str,
        value_col: str,
        agg: str = "sum",
        filter_col: str | None = None,
        filter_val: str | None = None,
        date_col: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> list[dict]:
        """Agrega as rows do dataset e retorna [{label, value}]."""
        computed = getattr(ds, 'computed_columns', None) or []
        computed_names = [c.get("name") for c in computed if c.get("name")]
        available = (ds.columns or []) + computed_names
        if agg not in self._VALID_AGGS:
            raise HTTPException(
                status_code=422,
                detail=f"Agregação inválida: '{agg}'. Use: {', '.join(sorted(self._VALID_AGGS))}",
            )
        if label_col not in available:
            raise HTTPException(
                status_code=422,
                detail=f"Coluna label '{label_col}' não encontrada. Disponíveis: {available}",
            )
        if value_col != "__count__" and value_col not in available:
            raise HTTPException(
                status_code=422,
                detail=f"Coluna value '{value_col}' não encontrada. Disponíveis: {available}",
            )
        if not ds.rows:
            return []
        rows = ds.rows
        if computed:
            rows = _apply_computed_columns(rows, computed)
        if filter_col and filter_val is not None:
            rows = [r for r in rows if str(r.get(filter_col, "")) == filter_val]
        if date_col and (date_from or date_to):
            d_from = _parse_date_value(date_from) if date_from else None
            d_to = _parse_date_value(date_to) if date_to else None
            filtered = []
            for r in rows:
                rd = _parse_date_value(r.get(date_col))
                if rd is None:
                    continue
                if d_from and rd < d_from:
                    continue
                if d_to and rd > d_to:
                    continue
                filtered.append(r)
            rows = filtered
        # Se value_col for "__count__", conta linhas por label
        if value_col == "__count__":
            groups: dict[str, int] = {}
            for row in rows:
                label = str(row.get(label_col, "")) or "(vazio)"
                groups[label] = groups.get(label, 0) + 1
            return [{"label": k, "value": v} for k, v in groups.items()]
        return _aggregate(rows, label_col, value_col, agg)

    async def add_computed_column(
        self,
        dataset_id: uuid.UUID,
        tenant_id: uuid.UUID,
        name: str,
        expression: str,
        refs: list[str],
    ) -> ReportDataset | None:
        ds = await self.get(dataset_id, tenant_id)
        if not ds:
            return None
        from sqlalchemy.orm.attributes import flag_modified
        cols = list(ds.computed_columns or [])
        # Remove existente com mesmo nome antes de inserir
        cols = [c for c in cols if c.get("name") != name]
        cols.append({"name": name, "expression": expression, "refs": refs})
        ds.computed_columns = cols
        flag_modified(ds, "computed_columns")
        await self.db.commit()
        await self.db.refresh(ds)
        return ds

    async def remove_computed_column(
        self,
        dataset_id: uuid.UUID,
        tenant_id: uuid.UUID,
        col_name: str,
    ) -> ReportDataset | None:
        ds = await self.get(dataset_id, tenant_id)
        if not ds:
            return None
        from sqlalchemy.orm.attributes import flag_modified
        cols = [c for c in (ds.computed_columns or []) if c.get("name") != col_name]
        ds.computed_columns = cols
        flag_modified(ds, "computed_columns")
        await self.db.commit()
        await self.db.refresh(ds)
        return ds

    async def _fetch_api(
        self, url: str, headers: dict, data_path: str | None
    ) -> tuple[list[dict], list[str]]:
        _validate_ssrf(url)

        _GS_HEADERS = {
            "User-Agent": "Mozilla/5.0 (compatible; Jarbis/1.0; +https://jarbis.cc)",
            "Accept": "text/csv,text/plain,*/*",
        }

        # Google Sheets: tenta gviz/tq primeiro (funciona com "qualquer pessoa com o link")
        # depois cai para export?format=csv — mesma lógica do endpoint preview-sheets
        import re as _re
        gs_match = _re.search(r"spreadsheets/d/([a-zA-Z0-9_-]+)", url)
        if gs_match:
            sheet_id = gs_match.group(1)
            sheet_m = _re.search(r"[?&]sheet=([^&]+)", url)
            gid_m = _re.search(r"[#&?]gid=(\d+)", url)
            if sheet_m:
                candidates = [
                    f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={sheet_m.group(1)}",
                    f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&sheet={sheet_m.group(1)}",
                ]
            else:
                gid = gid_m.group(1) if gid_m else "0"
                candidates = [
                    f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}",
                    f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}",
                ]
            content: bytes | None = None
            async with httpx.AsyncClient(follow_redirects=True, timeout=30) as client:
                for csv_url in candidates:
                    resp = await client.get(csv_url, headers=_GS_HEADERS)
                    if resp.status_code == 200:
                        ct = resp.headers.get("content-type", "")
                        if "text/html" in ct or resp.content[:15].lower().startswith(b"<!doctype"):
                            continue
                        content = resp.content
                        break
            if content is None:
                raise HTTPException(
                    status_code=400,
                    detail="Não foi possível acessar a planilha. Certifique-se de que está compartilhada como 'Qualquer pessoa com o link pode visualizar'.",
                )
            rows = _parse_csv(content)
            columns = _detect_columns(rows)
            return rows, columns

        # Outros tipos de API (JSON / CSV genérico)
        is_csv = "format=csv" in url

        async with httpx.AsyncClient(timeout=30, follow_redirects=False) as client:
            resp = await client.get(url, headers=headers)

        if resp.status_code in (301, 302, 303, 307, 308):
            redirect_url = resp.headers.get("location", "")
            if redirect_url:
                _validate_ssrf(redirect_url)  # valida destino do redirect antes de seguir
                async with httpx.AsyncClient(timeout=30, follow_redirects=False) as client2:
                    resp = await client2.get(redirect_url, headers=headers)

        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")
        if "text/csv" in content_type:
            is_csv = True

        if is_csv:
            rows = _parse_csv(resp.content)
        else:
            data = resp.json()
            raw_rows = _extract_json_path(data, data_path)
            rows = [_coerce_row(r) for r in raw_rows if isinstance(r, dict)]

        columns = _detect_columns(rows)
        return rows, columns
