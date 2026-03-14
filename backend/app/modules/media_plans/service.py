"""
Service — Planos de Mídia

Responsabilidades:
  - Parsear arquivos Excel (.xlsx / .xls) de planos de mídia
  - Detectar automaticamente abas de TV, Rádio, Jornal, OOH, Digital, etc.
  - Mapear colunas em português para os campos do modelo
  - Persistir MediaPlan + MediaPlanItems no banco
  - CRUD de planos por tenant
"""

import io
import json
import uuid
from datetime import date
from typing import Any

import openpyxl
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from .models import MediaPlan, MediaPlanItem

# ---------------------------------------------------------------------------
# Mapeamento de nomes de abas → meio
# ---------------------------------------------------------------------------

_ABA_MEIO: dict[str, str] = {
    "tv aberta": "TV",
    "tv": "TV",
    "televisao": "TV",
    "televisão": "TV",
    "grade tv": "TV",
    "grade": "TV",
    "pay tv": "TV",
    "tv paga": "TV",
    "radio": "RADIO",
    "rádio": "RADIO",
    "jornal": "JORNAL",
    "jornais": "JORNAL",
    "revista": "REVISTA",
    "revistas": "REVISTA",
    "ooh": "OOH",
    "out of home": "OOH",
    "mídia exterior": "OOH",
    "midia exterior": "OOH",
    "digital": "DIGITAL",
    "online": "DIGITAL",
    "internet": "DIGITAL",
    "performance": "DIGITAL",
    "awareness": "DIGITAL",
    "consideração": "DIGITAL",
    "consideracao": "DIGITAL",
    "social": "SOCIAL",
    "redes sociais": "SOCIAL",
    "social media": "SOCIAL",
    "influenciador": "INFLUENCIADOR",
    "influenciadores": "INFLUENCIADOR",
    "indoor": "INDOOR",
    "spotify": "DIGITAL",
    "streaming": "DIGITAL",
    "kpis": None,       # abas de resumo — ignorar
    "capa": None,
    "resumo": None,
    "cashflow": None,
    "budget": None,
    "cronograma": None,
    "flowchart": None,
    "comparativo": None,
    "cenários": None,
    "cenarios": None,
    "segmentações": None,
    "segmentacoes": None,
    "resultados": None,
}

# ---------------------------------------------------------------------------
# Mapeamento de colunas em português → campo do modelo
# ---------------------------------------------------------------------------

_COL_MAP: dict[str, str] = {
    # veículo
    "veiculo": "veiculo",
    "veículo": "veiculo",
    "emissora": "veiculo",
    "canal": "veiculo",
    "site": "veiculo",
    "portal": "veiculo",
    "programa": "programa",
    "formato": "formato",
    "posicionamento": "formato",
    "uf": "uf",
    "estado": "uf",
    "praça": "praca",
    "praca": "praca",
    "mercado": "praca",
    "municipio": "municipio",
    "município": "municipio",
    "cidade": "municipio",
    # período
    "inicio": "data_inicio",
    "início": "data_inicio",
    "data inicio": "data_inicio",
    "data início": "data_inicio",
    "data de início": "data_inicio",
    "fim": "data_fim",
    "data fim": "data_fim",
    "data de fim": "data_fim",
    "término": "data_fim",
    "termino": "data_fim",
    # financeiro
    "custo tabela": "custo_tabela",
    "custo de tabela": "custo_tabela",
    "valor tabela": "custo_tabela",
    "tabela": "custo_tabela",
    "desconto": "desconto_pct",
    "desconto %": "desconto_pct",
    "% desconto": "desconto_pct",
    "custo negociado": "custo_negociado",
    "valor negociado": "custo_negociado",
    "negociado": "custo_negociado",
    "custo total": "custo_total",
    "total": "custo_total",
    "investimento": "custo_total",
    "valor total": "custo_total",
    "bonificacao": "bonificacao",
    "bonificação": "bonificacao",
    # tv/radio
    "inserções": "insercoes",
    "insercoes": "insercoes",
    "spots": "insercoes",
    "grp": "grp",
    "trp": "trp",
    "audiencia": "audiencia_pct",
    "audiência": "audiencia_pct",
    "audiência %": "audiencia_pct",
    "rating": "audiencia_pct",
    "universo": "universo",
    "impactos": "impactos",
    "cpp": "cpp",
    "custo por ponto": "cpp",
    "cpm": "cpm",
    "custo por mil": "cpm",
    "alcance": "alcance_pct",
    "alcance %": "alcance_pct",
    "frequencia": "frequencia_media",
    "frequência": "frequencia_media",
    "frequência média": "frequencia_media",
    "opm": "ouvintes_por_minuto",
    "ouvintes por minuto": "ouvintes_por_minuto",
    # impresso
    "centimetragem": "centimetragem",
    "tiragem": "tiragem",
    "circulacao": "circulacao",
    "circulação": "circulacao",
    # digital
    "impressoes": "impressoes",
    "impressões": "impressoes",
    "cliques": "cliques",
    "clicks": "cliques",
    "ctr": "ctr_pct",
    "ctr %": "ctr_pct",
    "cpc": "cpc",
    "custo por clique": "cpc",
    "cpa": "cpa",
    "custo por aquisicao": "cpa",
    "custo por aquisição": "cpa",
    "conversoes": "conversoes",
    "conversões": "conversoes",
    "objetivo": "objetivo",
    # ooh
    "faces": "faces",
    "fluxo": "fluxo_diario",
    "fluxo diario": "fluxo_diario",
    "fluxo diário": "fluxo_diario",
}


def _normalize(s: Any) -> str:
    return str(s or "").strip().lower()


def _to_float(v: Any) -> float | None:
    if v is None or str(v).strip() in ("", "-", "n/a", "nd"):
        return None
    try:
        return float(str(v).replace(",", ".").replace("%", "").strip())
    except (ValueError, TypeError):
        return None


def _to_int(v: Any) -> int | None:
    f = _to_float(v)
    return int(f) if f is not None else None


def _to_date(v: Any) -> date | None:
    if isinstance(v, date):
        return v
    if v is None:
        return None
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%Y"):
        try:
            from datetime import datetime as _dt
            return _dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _detect_meio(sheet_name: str) -> str | None:
    """Retorna o meio detectado para a aba ou None se deve ser ignorada."""
    key = _normalize(sheet_name)
    # busca exata
    if key in _ABA_MEIO:
        return _ABA_MEIO[key]
    # busca parcial
    for token, meio in _ABA_MEIO.items():
        if token and token in key:
            return meio
    # se não mapeou mas parece útil, trata como "OUTRO"
    return "OUTRO"


def _parse_sheet(ws, meio: str, erros: list[str]) -> list[dict]:
    """Lê uma aba do Excel e retorna lista de dicts com os campos mapeados."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Encontra a linha de cabeçalho (primeira linha não vazia com texto)
    header_idx = None
    header_map: dict[int, str] = {}

    for row_idx, row in enumerate(rows):
        mapped = {}
        for col_idx, cell in enumerate(row):
            norm = _normalize(cell)
            if norm in _COL_MAP:
                mapped[col_idx] = _COL_MAP[norm]
        if len(mapped) >= 2:
            header_idx = row_idx
            header_map = mapped
            break

    if header_idx is None:
        return []

    items = []
    for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        # Linha vazia = pular
        non_empty = [c for c in row if c is not None and str(c).strip() not in ("", "-")]
        if len(non_empty) < 2:
            continue

        item: dict[str, Any] = {"meio": meio, "aba_origem": ws.title, "linha_origem": row_idx}

        for col_idx, field in header_map.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]

            if field in ("data_inicio", "data_fim"):
                item[field] = _to_date(val)
            elif field in (
                "custo_tabela", "desconto_pct", "custo_negociado", "custo_total",
                "bonificacao", "grp", "trp", "audiencia_pct", "cpp", "cpm",
                "alcance_pct", "frequencia_media", "centimetragem", "ctr_pct", "cpc", "cpa",
            ):
                item[field] = _to_float(val)
            elif field in (
                "insercoes", "impactos", "universo", "ouvintes_por_minuto",
                "tiragem", "circulacao", "impressoes", "cliques", "conversoes",
                "faces", "fluxo_diario",
            ):
                item[field] = _to_int(val)
            else:
                item[field] = str(val).strip() if val is not None else None

        # Precisa ter ao menos veículo ou custo para ser válido
        if not item.get("veiculo") and not item.get("custo_total") and not item.get("custo_negociado"):
            continue

        items.append(item)

    return items


def _extract_header_info(wb: openpyxl.Workbook) -> dict:
    """Tenta extrair metadados da aba 'Capa' ou primeira aba."""
    info: dict[str, Any] = {}
    capa = None
    for name in wb.sheetnames:
        if _normalize(name) in ("capa", "resumo", "dados gerais"):
            capa = wb[name]
            break
    if capa is None and wb.sheetnames:
        capa = wb[wb.sheetnames[0]]

    if capa:
        for row in capa.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                key = _normalize(cell)
                if key in ("anunciante", "cliente") and i + 1 < len(row) and row[i + 1]:
                    info["anunciante"] = str(row[i + 1]).strip()
                elif key in ("agencia", "agência") and i + 1 < len(row) and row[i + 1]:
                    info["agencia"] = str(row[i + 1]).strip()
                elif key == "produto" and i + 1 < len(row) and row[i + 1]:
                    info["produto"] = str(row[i + 1]).strip()
                elif key in ("campanha", "plano") and i + 1 < len(row) and row[i + 1]:
                    info.setdefault("nome", str(row[i + 1]).strip())
                elif key in ("periodo", "período", "início", "inicio") and i + 1 < len(row) and row[i + 1]:
                    info.setdefault("data_inicio", _to_date(row[i + 1]))
                elif key in ("fim", "término", "termino") and i + 1 < len(row) and row[i + 1]:
                    info.setdefault("data_fim", _to_date(row[i + 1]))
    return info


class MediaPlanService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def upload_xlsx(
        self,
        tenant_id: uuid.UUID,
        filename: str,
        file_bytes: bytes,
    ) -> MediaPlan:
        """Parse o arquivo Excel e persiste o plano no banco."""
        erros: list[str] = []
        all_items: list[dict] = []
        header_info: dict = {}

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes), read_only=True, data_only=True
            )
            header_info = _extract_header_info(wb)

            for sheet_name in wb.sheetnames:
                meio = _detect_meio(sheet_name)
                if meio is None:
                    continue  # aba de resumo/capa, ignora
                try:
                    ws = wb[sheet_name]
                    items = _parse_sheet(ws, meio, erros)
                    all_items.extend(items)
                except Exception as e:
                    erros.append(f"Aba '{sheet_name}': {e}")

        except Exception as e:
            erros.append(f"Erro ao abrir arquivo: {e}")

        # Define nome do plano
        nome = header_info.get("nome") or filename.replace(".xlsx", "").replace(".xls", "")

        # Calcula budget_total como soma dos custo_total dos itens
        budget_total = sum(
            it.get("custo_total") or it.get("custo_negociado") or 0
            for it in all_items
            if (it.get("custo_total") or it.get("custo_negociado"))
        ) or None

        parse_status = "OK" if not erros else ("PARCIAL" if all_items else "ERRO")

        plan = MediaPlan(
            tenant_id=tenant_id,
            nome=nome,
            anunciante=header_info.get("anunciante"),
            agencia=header_info.get("agencia"),
            produto=header_info.get("produto"),
            data_inicio=header_info.get("data_inicio"),
            data_fim=header_info.get("data_fim"),
            budget_total=budget_total,
            arquivo_nome=filename,
            arquivo_tamanho_bytes=len(file_bytes),
            parse_status=parse_status,
            parse_erros=json.dumps(erros, ensure_ascii=False) if erros else None,
        )
        self.db.add(plan)
        await self.db.flush()  # gera plan.id antes de criar items

        for it in all_items:
            item = MediaPlanItem(
                plan_id=plan.id,
                tenant_id=tenant_id,
                **it,
            )
            self.db.add(item)

        await self.db.commit()
        await self.db.refresh(plan)
        return plan

    async def list_plans(self, tenant_id: uuid.UUID) -> list[dict]:
        """Lista resumo dos planos do tenant com totais calculados."""
        result = await self.db.execute(
            select(
                MediaPlan,
                func.count(MediaPlanItem.id).label("total_items"),
                func.sum(
                    func.coalesce(MediaPlanItem.custo_total, MediaPlanItem.custo_negociado, 0)
                ).label("investimento_total"),
            )
            .outerjoin(MediaPlanItem, MediaPlanItem.plan_id == MediaPlan.id)
            .where(MediaPlan.tenant_id == tenant_id)
            .group_by(MediaPlan.id)
            .order_by(MediaPlan.created_at.desc())
        )
        rows = result.all()
        plans = []
        for plan, total_items, investimento in rows:
            plans.append({
                "id": plan.id,
                "nome": plan.nome,
                "anunciante": plan.anunciante,
                "agencia": plan.agencia,
                "data_inicio": plan.data_inicio,
                "data_fim": plan.data_fim,
                "budget_total": plan.budget_total,
                "parse_status": plan.parse_status,
                "arquivo_nome": plan.arquivo_nome,
                "created_at": plan.created_at,
                "total_items": total_items or 0,
                "investimento_total": float(investimento) if investimento else None,
            })
        return plans

    async def get_plan(self, plan_id: uuid.UUID, tenant_id: uuid.UUID) -> MediaPlan | None:
        result = await self.db.execute(
            select(MediaPlan)
            .options(selectinload(MediaPlan.items))
            .where(MediaPlan.id == plan_id, MediaPlan.tenant_id == tenant_id)
        )
        return result.scalar_one_or_none()

    async def delete_plan(self, plan_id: uuid.UUID, tenant_id: uuid.UUID) -> bool:
        plan = await self.get_plan(plan_id, tenant_id)
        if not plan:
            return False
        await self.db.delete(plan)
        await self.db.commit()
        return True

    async def list_material_items(self, tenant_id: uuid.UUID) -> list[dict]:
        """Retorna todos os itens com dados de material do tenant, enriquecidos com info do plano."""
        result = await self.db.execute(
            select(MediaPlanItem, MediaPlan.nome, MediaPlan.anunciante)
            .join(MediaPlan, MediaPlan.id == MediaPlanItem.plan_id)
            .where(MediaPlanItem.tenant_id == tenant_id)
            .order_by(MediaPlanItem.material_prazo_entrega.asc().nullslast(), MediaPlan.nome)
        )
        rows = result.all()
        items = []
        for item, plan_nome, anunciante in rows:
            items.append({
                "id": item.id,
                "plan_id": item.plan_id,
                "plan_nome": plan_nome,
                "anunciante": anunciante,
                "meio": item.meio,
                "veiculo": item.veiculo,
                "formato": item.formato,
                "material_formato": item.material_formato,
                "material_dimensoes": item.material_dimensoes,
                "material_tipo_fechamento": item.material_tipo_fechamento,
                "material_prazo_entrega": item.material_prazo_entrega,
                "material_horario_entrega": item.material_horario_entrega,
                "material_status": item.material_status,
                "material_entregue_em": item.material_entregue_em,
                "observacoes": item.observacoes,
            })
        return items

    async def update_item(
        self, plan_id: uuid.UUID, item_id: uuid.UUID, tenant_id: uuid.UUID, data: dict
    ) -> MediaPlanItem | None:
        result = await self.db.execute(
            select(MediaPlanItem).where(
                MediaPlanItem.id == item_id,
                MediaPlanItem.plan_id == plan_id,
                MediaPlanItem.tenant_id == tenant_id,
            )
        )
        item = result.scalar_one_or_none()
        if not item:
            return None
        for k, v in data.items():
            setattr(item, k, v)
        await self.db.commit()
        await self.db.refresh(item)
        return item
